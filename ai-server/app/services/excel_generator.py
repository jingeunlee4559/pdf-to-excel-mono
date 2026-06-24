from __future__ import annotations

import json
import os
import re
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet



def request_output_intent(text: str = "") -> str:
    raw = str(text or "")
    wants_report = any(token in raw for token in ["보고서 형식", "보고서", "업무보고서", "검토보고서", "서술형", "문장형", "보고용"])
    wants_table = any(token in raw for token in ["표로", "표 형태", "표 형식", "비교표", "단가표", "조사표", "테이블", "그리드", "엑셀 표"])
    if wants_table and not wants_report:
        return "TABLE"
    if wants_table and re.search(r"표로|비교표|단가표|조사표|테이블|그리드|엑셀\s*표", raw):
        return "TABLE"
    if wants_report:
        return "REPORT"
    return "AUTO"

BASE_DIR = Path(__file__).resolve().parents[2]
RESULT_DIR = BASE_DIR / "storage" / "results"
TEMPLATE_DIR = BASE_DIR / "storage" / "templates" / "ai_generated"

THIN = Side(style="thin", color="3B4A5A")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="DDE6F1")
LIGHT_FILL = PatternFill("solid", fgColor="EEF7FF")
TITLE_FILL = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="E9F8EF")
AMBER_FILL = PatternFill("solid", fgColor="FFF4D6")

DOCUMENT_TYPE_LABELS = {
    "ESTIMATE": "견적서",
    "ESTIMATE_COMPARISON": "비교 견적서",
    "MULTI_VENDOR_PRICE_COMPARISON": "비교 견적서",
    "UNIT_PRICE_TABLE": "단가표",
    "STANDARD_MARKET_PRICE_TABLE": "단가표",
    "REPORT": "보고서",
    "BUSINESS_REPORT": "보고서",
    "MEETING_MINUTES": "회의록",
    "OFFICIAL_LETTER": "공문",
    "NORMAL_TABLE": "문서 정리표",
    "CUSTOM_DOCUMENT_FORM": "문서 양식",
}

TYPE_KEYWORDS = {
    "ESTIMATE_COMPARISON": ["비교", "견적", "업체", "회사", "단가비교", "가격비교"],
    "UNIT_PRICE_TABLE": ["단가표", "단가", "표준시장", "품셈", "가격표"],
    "REPORT": ["보고서", "보고", "현황", "요약", "검토"],
    "MEETING_MINUTES": ["회의록", "회의", "안건", "참석", "결정사항"],
    "OFFICIAL_LETTER": ["공문", "수신", "참조", "제목", "시행"],
}

BASE_LABELS = {
    "row_no": "NO",
    "no": "NO",
    "project_name": "공사명",
    "site_name": "현장명",
    "document_title": "제목",
    "document_date": "작성일",
    "requester_name": "작성자",
    "writer_name": "작성자",
    "vendor_name": "업체명",
    "company_name": "회사명",
    "construction_code": "공종코드",
    "item_name": "품명",
    "product_name": "품명",
    "work_item_name": "공종명칭",
    "spec": "규격",
    "standard": "규격",
    "quantity": "수량",
    "unit": "단위",
    "unit_price": "단가",
    "standard_unit_price": "표준단가",
    "vendor_unit_price": "업체단가",
    "amount": "금액",
    "total_amount": "합계",
    "lowest_target": "최저 업체",
    "lowest_vendor": "최저 업체",
    "calculated_unit_price": "최저 단가",
    "lowest_unit_price": "최저 단가",
    "remark": "비고",
    "note": "비고",
    "meeting_date": "회의일자",
    "attendees": "참석자",
    "agenda": "안건",
    "decision": "결정사항",
    "action_item": "조치사항",
    "recipient": "수신",
    "reference": "참조",
    "sender": "발신",
    "body": "내용",
    "summary": "요약",
    "opinion": "검토의견",
}


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def safe_filename(name: Optional[str], fallback: str = "document_result.xlsx") -> str:
    raw = str(name or fallback).strip() or fallback
    raw = re.sub(r'[\\/:*?"<>|]+', "_", raw)
    if not raw.lower().endswith(".xlsx"):
        raw += ".xlsx"
    return raw


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def today_text() -> str:
    return datetime.now().strftime("%Y.%m.%d")


def as_list(value: Any) -> List[Any]:
    return value if isinstance(value, list) else []


def normalize_key(value: Any) -> str:
    return str(value or "").strip()


def label_for(key: str, fallback: Optional[str] = None) -> str:
    return str(fallback or BASE_LABELS.get(key) or key)


def to_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    raw = re.sub(r"[^0-9.\-]", "", str(value or ""))
    try:
        return float(raw) if raw not in ("", "-", ".") else 0.0
    except Exception:
        return 0.0


def write_cell(ws: Worksheet, row: int, col: int, value: Any, *, bold: bool = False, fill: Optional[PatternFill] = None,
               align: str = "center", border: bool = True, wrap: bool = True) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    if border:
        cell.border = BORDER


def merge_write(ws: Worksheet, start_row: int, start_col: int, end_row: int, end_col: int, value: Any,
                *, bold: bool = False, fill: Optional[PatternFill] = None, size: int = 11) -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = ws.cell(start_row, start_col, value)
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).border = BORDER


def set_widths(ws: Worksheet, count: int, base: int = 14) -> None:
    for c in range(1, count + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(base, min(28, base + (c % 3)))


def get_row_value(row: Dict[str, Any], key: str, index: int = 0) -> Any:
    aliases = {
        "row_no": ["row_no", "no", "순번"],
        "item_name": ["item_name", "product_name", "work_item_name", "품명", "품목", "공종명칭", "항목"],
        "spec": ["spec", "standard", "규격", "사양"],
        "quantity": ["quantity", "qty", "request_quantity", "requested_quantity", "수량"],
        "unit": ["unit", "단위"],
        "unit_price": ["unit_price", "vendor_unit_price", "price", "단가"],
        "amount": ["amount", "total_amount", "금액"],
        "remark": ["remark", "note", "memo", "비고"],
    }
    if key in ("row_no", "no"):
        return row.get("row_no") or row.get("no") or index + 1
    for alias in aliases.get(key, [key]):
        if alias in row and row.get(alias) not in (None, ""):
            return row.get(alias)
    return row.get(key, "")



def first_text(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        if isinstance(value, list):
            parts = []
            for item in value:
                if isinstance(item, dict):
                    parts.append(str(item.get("value") or item.get("content") or item.get("text") or ""))
                else:
                    parts.append(str(item or ""))
            text = "\n".join(part.strip() for part in parts if part and str(part).strip()).strip()
            if text:
                return text
            continue
        if isinstance(value, dict):
            text = str(value.get("value") or value.get("content") or value.get("text") or value.get("summary") or "").strip()
            if text:
                return text
            continue
        text = str(value or "").strip()
        if text:
            return text
    return ""


def compact_status(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_json_maybe(value: Any) -> Any:
    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return {}
    return value if isinstance(value, dict) else {}


def get_payload_analysis(payload: Dict[str, Any]) -> Dict[str, Any]:
    job = parse_json_maybe(payload.get("job"))
    analysis = parse_json_maybe(job.get("analysis"))
    raw = parse_json_maybe(analysis.get("raw"))
    merged = {**raw, **analysis}
    return merged


def get_payload_drafts(payload: Dict[str, Any]) -> Dict[str, Any]:
    analysis = get_payload_analysis(payload)
    drafts = analysis.get("drafts") if isinstance(analysis.get("drafts"), dict) else {}
    return drafts


def ignore_plain_status(value: Any) -> bool:
    text = compact_status(value)
    return not text or text in {"정상", "완료", "없음", "해당없음", "확인"} or len(text) <= 2


def looks_like_user_prompt_or_meta(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    prompt_tokens = [
        "정리해줘", "작성해줘", "만들어줘", "써줘", "출력해줘", "보여줘",
        "보고서 형식", "보고서 형태", "보고서로", "요약과 검토내용", "상세하게", "풍부하게",
    ]
    meta_pattern = r"첨부\s*파일|총\s*페이지|확인된\s*총\s*페이지|표\s*후보|PyMuPDF|pdfplumber|PP-Structure|PaddleOCR|OCR|LLM|ai-server|저장\s*위치|산출\s*방식|요청\s*내용은|백그라운드|엑셀\s*미리보기"
    return any(token in text for token in prompt_tokens) or bool(re.search(meta_pattern, text, re.I))


def document_only_text(value: Any) -> str:
    text = str(value or "").strip()
    if not text or looks_like_user_prompt_or_meta(text):
        return ""
    lines = [line.strip() for line in re.split(r"\r?\n", text) if line.strip()]
    clean_lines = [line for line in lines if not looks_like_user_prompt_or_meta(line)]
    return "\n".join(clean_lines)


def build_report_first_row(payload: Dict[str, Any], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    first = rows[0] if rows else {}
    analysis = get_payload_analysis(payload)
    report = get_payload_drafts(payload).get("report") if isinstance(get_payload_drafts(payload).get("report"), dict) else {}
    issue_value = first_text(first.get("issue_summary"), first.get("review_result"), first.get("review_opinion"))
    if ignore_plain_status(issue_value):
        issue_value = first_text(report.get("issue_summary"), report.get("review_result"), report.get("review_opinion"))
    return {
        **first,
        "report_title": first_text(document_only_text(first.get("report_title")), document_only_text(first.get("document_title")), document_only_text(report.get("report_title")), document_only_text(report.get("title")), "업무 보고서"),
        "report_purpose": first_text(document_only_text(first.get("report_purpose")), document_only_text(first.get("purpose")), document_only_text(report.get("report_purpose")), document_only_text(report.get("purpose")), document_only_text(analysis.get("purpose")), "첨부 문서의 주요 내용을 업무 보고서 형식으로 정리합니다."),
        "summary": first_text(document_only_text(first.get("summary")), document_only_text(first.get("content")), document_only_text(report.get("summary")), document_only_text(analysis.get("summary"))),
        "issue_summary": first_text(document_only_text(issue_value)),
        "action_plan": first_text(document_only_text(first.get("action_plan")), document_only_text(report.get("action_plan"))),
        "footer_note": first_text(document_only_text(first.get("footer_note")), document_only_text(report.get("footer_note"))),
    }


def is_explicit_action_text(value: Any) -> bool:
    text = compact_status(value)
    if len(text) < 6:
        return False
    if re.fullmatch(r"(물량|일정|품질|안전|회의|작업일보|협력업체)", text):
        return False
    return bool(re.search(r"조치|처리|확인|점검|보완|검토|지시|미종결|변경|협의|관리|요청|제출|작성|보고|수립", text))


def build_meeting_rows(payload: Dict[str, Any], rows: List[Dict[str, Any]]) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    analysis = get_payload_analysis(payload)
    meeting = get_payload_drafts(payload).get("meeting") if isinstance(get_payload_drafts(payload).get("meeting"), dict) else {}
    first = rows[0] if rows else {}
    base = {
        **first,
        "meeting_title": first_text(first.get("meeting_title"), first.get("document_title"), meeting.get("meeting_title"), meeting.get("title"), "업무 회의록"),
        "meeting_date": first_text(first.get("meeting_date"), meeting.get("meeting_date")),
        "meeting_place": first_text(first.get("meeting_place"), meeting.get("meeting_place")),
        "attendees": first_text(first.get("attendees"), meeting.get("attendees")),
        "agenda": first_text(first.get("agenda"), meeting.get("agenda"), analysis.get("purpose"), "첨부 문서 검토 및 후속 관리 방안 논의"),
        "discussion": first_text(first.get("discussion"), first.get("content"), meeting.get("discussion"), analysis.get("summary")),
        "decision": first_text(first.get("decision"), meeting.get("decision"), "원문에 명시된 최종 결정사항은 확인되지 않았습니다. 회의 확정 후 입력하세요."),
        "remark": first_text(first.get("remark"), meeting.get("remark"), "문서 분석 결과 기준 회의록 초안입니다."),
    }
    row_actions = []
    for row in rows:
        action = first_text(row.get("action_item"), row.get("todo"), row.get("next_action"))
        if is_explicit_action_text(action):
            row_actions.append({**row, "action_item": action})
    draft_actions = meeting.get("action_items") if isinstance(meeting.get("action_items"), list) else []
    if not row_actions:
        for item in draft_actions:
            if not isinstance(item, dict):
                continue
            action = first_text(item.get("action_item"), item.get("content"), item.get("text"))
            if is_explicit_action_text(action):
                row_actions.append({
                    "action_item": action,
                    "owner": first_text(item.get("owner"), item.get("assignee"), "확인 필요"),
                    "due_date": first_text(item.get("due_date"), item.get("dueDate"), "미정"),
                    "status": first_text(item.get("status"), "확인 필요"),
                    "remark": first_text(item.get("remark"), item.get("note")),
                })
    if not row_actions:
        row_actions = [{
            "action_item": "원문에 명시된 후속 조치사항은 확인되지 않았습니다. 필요 시 담당자와 기한을 지정하세요.",
            "owner": "확인 필요",
            "due_date": "미정",
            "status": "확인 필요",
            "remark": "",
        }]
    return base, row_actions


def build_official_first_row(payload: Dict[str, Any], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    first = rows[0] if rows else {}
    analysis = get_payload_analysis(payload)
    official = get_payload_drafts(payload).get("officialLetter") if isinstance(get_payload_drafts(payload).get("officialLetter"), dict) else {}
    return {
        **first,
        "letter_title": first_text(first.get("letter_title"), official.get("letter_title"), "공 문"),
        "document_no": first_text(first.get("document_no"), official.get("document_no")),
        "recipient": first_text(first.get("recipient"), official.get("recipient"), "수신처 확인 필요"),
        "reference": first_text(first.get("reference"), official.get("reference")),
        "document_title": first_text(first.get("document_title"), first.get("title"), official.get("document_title"), analysis.get("purpose"), "업무 협조 요청"),
        "body": first_text(first.get("body"), first.get("content"), official.get("body"), analysis.get("summary")),
        "attachment_note": first_text(first.get("attachment_note"), official.get("attachment_note"), "첨부 문서 참조"),
        "sender": first_text(first.get("sender"), official.get("sender"), payload.get("author_name"), "공사팀"),
    }


def comparable_company(value: Any) -> str:
    text = str(value or "")
    text = re.sub(r"주식회사|\(주\)|㈜|（주）", "", text)
    return re.sub(r"[\s._\-()（）\[\]{}·,]", "", text).lower()


def normalize_vendor_label(label: Any) -> str:
    return re.sub(r"\s*(단가|금액|견적가|견적단가|가격)$", "", str(label or "")).strip()


def infer_vendors(columns: List[Dict[str, Any]], rows: List[Dict[str, Any]], table_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    vendor_map: Dict[str, Dict[str, Any]] = {}

    def put(name: Any, patch: Optional[Dict[str, Any]] = None) -> Optional[Dict[str, Any]]:
        display = str(name or "").strip()
        key = comparable_company(display)
        if not display or not key or re.match(r"^(기준|표준|일반|최저|차이|수량|단가|금액|품명|규격)$", display):
            return None
        current = vendor_map.get(key, {"name": display, "compareKey": key, "index": len(vendor_map)})
        current.update(patch or {})
        current["name"] = current.get("name") or display
        vendor_map[key] = current
        return current

    for idx, vendor in enumerate(as_list((table_json or {}).get("meta", {}).get("vendors"))):
        if isinstance(vendor, dict):
            name = vendor.get("name") or vendor.get("vendorName") or vendor.get("label")
            put(name, {
                "index": idx,
                "nameKey": vendor.get("nameKey"),
                "unitPriceKey": vendor.get("unitPriceKey") or vendor.get("priceKey"),
                "amountKey": vendor.get("amountKey"),
                "quantityKey": vendor.get("quantityKey"),
            })
        else:
            put(vendor, {"index": idx})

    for col in columns or []:
        key = str(col.get("key") or "")
        label = str(col.get("label") or key)
        m = re.match(r"^(?:vendor|company|target)[_\-]?(\d+)[_\-]?(name|spec|quantity|qty|unit_price|price|amount)$", key, re.I)
        if m:
            raw_idx = int(m.group(1))
            zero_idx = raw_idx - 1 if raw_idx > 0 else raw_idx
            field = m.group(2).lower()
            name = ""
            for row in rows or []:
                name = row.get(f"vendor_{raw_idx}_name") or row.get(f"company_{raw_idx}_name") or ""
                if name:
                    break
            name = name or normalize_vendor_label(label)
            vendor = put(name, {"index": zero_idx})
            if not vendor:
                continue
            if field == "name": vendor["nameKey"] = key
            if field in ("unit_price", "price"): vendor["unitPriceKey"] = key
            if field == "amount": vendor["amountKey"] = key
            if field in ("quantity", "qty"): vendor["quantityKey"] = key
            continue
        if re.search(r"(단가|금액|견적가|견적단가|가격)$", label):
            name = normalize_vendor_label(label)
            vendor = put(name)
            if vendor:
                if label.endswith("금액"):
                    vendor["amountKey"] = key
                else:
                    vendor["unitPriceKey"] = key

    if not vendor_map:
        names = []
        for row in rows or []:
            name = str(row.get("vendor_name") or row.get("target_name") or row.get("company_name") or "").strip()
            if name and name not in names:
                names.append(name)
        for idx, name in enumerate(names):
            put(name, {"index": idx, "unitPriceKey": "vendor_unit_price", "amountKey": "amount"})

    return sorted(vendor_map.values(), key=lambda v: v.get("index", 999))


def get_vendor_value(row: Dict[str, Any], vendor: Dict[str, Any], field_key: str) -> Any:
    if field_key in ("target_name", "vendor_name", "company_name"):
        return vendor.get("name", "")
    if field_key in ("quantity", "qty"):
        return row.get(vendor.get("quantityKey") or "") or get_row_value(row, "quantity")
    if field_key in ("unit_price", "vendor_unit_price", "price"):
        key = vendor.get("unitPriceKey")
        if key and row.get(key) not in (None, ""):
            return row.get(key)
        maps = row.get("vendor_prices") or row.get("vendorPrices") or row.get("vendor_unit_prices") or {}
        if isinstance(maps, dict):
            for k, v in maps.items():
                if comparable_company(k) == comparable_company(vendor.get("name")):
                    return v
        if comparable_company(row.get("vendor_name")) == comparable_company(vendor.get("name")):
            return row.get("vendor_unit_price") or row.get("unit_price") or ""
        return row.get("vendor_unit_price") if not row.get("vendor_name") else ""
    if field_key in ("amount", "vendor_amount"):
        key = vendor.get("amountKey")
        if key and row.get(key) not in (None, ""):
            return row.get(key)
        maps = row.get("vendor_amounts") or row.get("vendorAmounts") or {}
        if isinstance(maps, dict):
            for k, v in maps.items():
                if comparable_company(k) == comparable_company(vendor.get("name")):
                    return v
        qty = to_number(get_vendor_value(row, vendor, "quantity"))
        price = to_number(get_vendor_value(row, vendor, "unit_price"))
        return int(qty * price) if qty and price else ""
    return row.get(field_key, "")


def lowest_vendor(row: Dict[str, Any], vendors: List[Dict[str, Any]]) -> Tuple[str, Any]:
    best_name = ""
    best_price = 0.0
    for vendor in vendors:
        price = to_number(get_vendor_value(row, vendor, "unit_price"))
        if price and (not best_price or price < best_price):
            best_name = vendor.get("name", "")
            best_price = price
    return best_name, int(best_price) if best_price else ""


def detect_document_type(payload: Dict[str, Any]) -> str:
    mapping = payload.get("mapping_json") or {}
    template = payload.get("template") or {}
    job = payload.get("job") or {}
    analysis = job.get("analysis") or {}
    table_type = str((payload.get("job") or {}).get("tables", [{}])[0].get("tableType", "") if isinstance((payload.get("job") or {}).get("tables"), list) else "")
    text = " ".join(str(x or "") for x in [
        payload.get("output_mode"), mapping.get("template_type"), mapping.get("templateType"), mapping.get("layout"),
        template.get("template_type"), template.get("templateType"), template.get("template_name"), template.get("templateName"),
        job.get("userRequest"), job.get("user_request"), analysis.get("documentType"), analysis.get("recommendedTableType"), table_type,
    ]).upper()
    korean_text = " ".join(str(x or "") for x in [template.get("template_name"), template.get("templateName"), job.get("userRequest"), job.get("user_request")])
    if "MULTI_VENDOR" in text or "COMPARISON" in text or any(k in korean_text for k in TYPE_KEYWORDS["ESTIMATE_COMPARISON"]):
        return "ESTIMATE_COMPARISON"
    if "UNIT_PRICE" in text or "STANDARD_MARKET" in text or any(k in korean_text for k in TYPE_KEYWORDS["UNIT_PRICE_TABLE"]):
        return "UNIT_PRICE_TABLE"
    if "CUSTOM_DOCUMENT_FORM" in text or "DOCUMENT_FORM" in text:
        return "CUSTOM_DOCUMENT_FORM"
    if "MEETING" in text or any(k in korean_text for k in TYPE_KEYWORDS["MEETING_MINUTES"]):
        return "MEETING_MINUTES"
    if "OFFICIAL" in text or any(k in korean_text for k in TYPE_KEYWORDS["OFFICIAL_LETTER"]):
        return "OFFICIAL_LETTER"
    if "REPORT" in text or any(k in korean_text for k in TYPE_KEYWORDS["REPORT"]):
        return "REPORT"
    return "NORMAL_TABLE"


def normalize_columns(columns: List[Dict[str, Any]], rows: List[Dict[str, Any]], doc_type: str, mapping_json: Optional[Dict[str, Any]] = None) -> List[Dict[str, str]]:
    mapping_json = mapping_json or {}
    if mapping_json.get("baseColumns"):
        out = []
        for item in as_list(mapping_json.get("baseColumns")):
            key = normalize_key(item.get("fieldKey") or item.get("key"))
            if key:
                out.append({"key": key, "label": label_for(key, item.get("label"))})
        if not any(c["key"] in ("row_no", "no") for c in out):
            out.insert(0, {"key": "row_no", "label": "NO"})
        return out

    seen = set()
    out = []
    preferred = {
        "ESTIMATE_COMPARISON": ["row_no", "item_name", "spec", "quantity", "unit"],
        "UNIT_PRICE_TABLE": ["row_no", "construction_code", "item_name", "spec", "unit", "unit_price", "amount", "remark"],
        "REPORT": ["row_no", "section", "summary", "item_name", "content", "remark"],
        "MEETING_MINUTES": ["row_no", "agenda", "decision", "action_item", "owner", "due_date", "remark"],
        "OFFICIAL_LETTER": ["row_no", "recipient", "reference", "document_title", "body", "remark"],
        "NORMAL_TABLE": ["row_no", "item_name", "spec", "quantity", "unit", "unit_price", "amount", "remark"],
    }.get(doc_type, [])

    def add(key: str, label: Optional[str] = None) -> None:
        if not key or key in seen:
            return
        seen.add(key)
        out.append({"key": key, "label": label_for(key, label)})

    for key in preferred:
        add(key)
    for col in columns or []:
        key = normalize_key(col.get("key") or col.get("fieldKey"))
        if re.match(r"^(vendor|company|target)_?\d+_", key, re.I):
            continue
        add(key, col.get("label"))
    for row in rows[:10] if rows else []:
        for key, val in row.items():
            if val not in (None, "") and not re.match(r"^(vendor|company|target)_?\d+_", key, re.I):
                add(key)
    return out[:18]


def write_title_area(ws: Worksheet, title: str, total_cols: int, subtitle: str = "") -> None:
    ws.row_dimensions[1].height = 30
    merge_write(ws, 1, 1, 1, total_cols, title, bold=True, size=16, fill=TITLE_FILL)
    if subtitle:
        merge_write(ws, 2, 1, 2, total_cols, subtitle, bold=False, size=10, fill=LIGHT_FILL)
    write_cell(ws, 3, 1, "작성일", bold=True, fill=HEADER_FILL)
    write_cell(ws, 3, 2, today_text())
    write_cell(ws, 3, 3, "작성자", bold=True, fill=HEADER_FILL)
    write_cell(ws, 3, 4, "")


def write_table(ws: Worksheet, columns: List[Dict[str, str]], rows: List[Dict[str, Any]], start_row: int = 5) -> None:
    for idx, col in enumerate(columns, start=1):
        write_cell(ws, start_row, idx, col.get("label") or col.get("key"), bold=True, fill=HEADER_FILL)
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, col in enumerate(columns, start=1):
            key = col.get("key")
            value = get_row_value(row, key, r_idx - start_row)
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left" if key in ("item_name", "summary", "content", "body", "remark") else "center", vertical="center", wrap_text=True)
            if re.search(r"price|amount|cost|total|quantity|수량|단가|금액", str(key), re.I):
                num = to_number(value)
                if value not in (None, "") and num:
                    cell.value = int(num) if float(num).is_integer() else num
                    cell.number_format = "#,##0"
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(columns))}{start_row + len(rows)}"
    ws.freeze_panes = ws.cell(start_row + 1, 1)
    set_widths(ws, len(columns))


def create_free_form_workbook(payload: Dict[str, Any]) -> Tuple[Workbook, Dict[str, Any]]:
    rows = as_list(payload.get("rows"))
    columns = as_list(payload.get("columns"))
    doc_type = detect_document_type(payload)
    wb = Workbook()
    ws = wb.active
    ws.title = DOCUMENT_TYPE_LABELS.get(doc_type, "문서정리")[:31]
    normalized_columns = normalize_columns(columns, rows, doc_type, payload.get("mapping_json") or {})
    title = DOCUMENT_TYPE_LABELS.get(doc_type, "문서 정리표")
    write_title_area(ws, title, max(len(normalized_columns), 6), "분석된 표 데이터를 기준으로 생성한 자유형 엑셀입니다.")
    write_table(ws, normalized_columns, rows, 5)
    return wb, {"template_kind": doc_type, "columns": normalized_columns}


def create_estimate_comparison_workbook(payload: Dict[str, Any]) -> Tuple[Workbook, Dict[str, Any]]:
    rows = as_list(payload.get("rows"))
    columns = as_list(payload.get("columns"))
    job = payload.get("job") or {}
    table_json = {}
    if isinstance(job.get("tables"), list) and job["tables"]:
        table_json = job["tables"][0].get("tableJson") or job["tables"][0].get("table_json") or {}
    vendors = infer_vendors(columns, rows, table_json)
    if not vendors:
        vendors = [{"name": "업체1", "index": 0}, {"name": "업체2", "index": 1}]
    base_cols = [
        {"key": "row_no", "label": "NO"}, {"key": "item_name", "label": "품명"},
        {"key": "spec", "label": "규격"}, {"key": "quantity", "label": "수량"}, {"key": "unit", "label": "단위"},
    ]
    total_cols = len(base_cols) + len(vendors) * 2 + 3
    wb = Workbook()
    ws = wb.active
    ws.title = "비교견적서"
    write_title_area(ws, "비교 견적서", total_cols, "업체별 단가와 금액을 비교합니다.")
    header_row = 5
    for idx, col in enumerate(base_cols, start=1):
        merge_write(ws, header_row, idx, header_row + 1, idx, col["label"], bold=True, fill=HEADER_FILL)
    col_cursor = len(base_cols) + 1
    for vendor in vendors:
        merge_write(ws, header_row, col_cursor, header_row, col_cursor + 1, vendor.get("name") or "업체", bold=True, fill=HEADER_FILL)
        write_cell(ws, header_row + 1, col_cursor, "단가", bold=True, fill=LIGHT_FILL)
        write_cell(ws, header_row + 1, col_cursor + 1, "금액", bold=True, fill=LIGHT_FILL)
        col_cursor += 2
    for label in ["최저 업체", "최저 단가", "비고"]:
        merge_write(ws, header_row, col_cursor, header_row + 1, col_cursor, label, bold=True, fill=HEADER_FILL)
        col_cursor += 1

    for row_offset, row in enumerate(rows):
        r_idx = header_row + 2 + row_offset
        for c_idx, col in enumerate(base_cols, start=1):
            val = get_row_value(row, col["key"], row_offset)
            write_cell(ws, r_idx, c_idx, val, align="left" if col["key"] == "item_name" else "center")
        c = len(base_cols) + 1
        best_name, best_price = lowest_vendor(row, vendors)
        for vendor in vendors:
            price = get_vendor_value(row, vendor, "unit_price")
            amount = get_vendor_value(row, vendor, "amount")
            write_cell(ws, r_idx, c, int(to_number(price)) if price not in (None, "") and to_number(price) else price)
            ws.cell(r_idx, c).number_format = "#,##0"
            write_cell(ws, r_idx, c + 1, int(to_number(amount)) if amount not in (None, "") and to_number(amount) else amount)
            ws.cell(r_idx, c + 1).number_format = "#,##0"
            c += 2
        write_cell(ws, r_idx, c, best_name)
        write_cell(ws, r_idx, c + 1, best_price)
        ws.cell(r_idx, c + 1).number_format = "#,##0"
        write_cell(ws, r_idx, c + 2, get_row_value(row, "remark"), align="left")
    ws.freeze_panes = ws.cell(header_row + 2, 1)
    set_widths(ws, total_cols, 13)
    return wb, {"template_kind": "ESTIMATE_COMPARISON", "vendor_count": len(vendors)}



def create_estimate_form_workbook(payload: Dict[str, Any]) -> Tuple[Workbook, Dict[str, Any]]:
    rows = as_list(payload.get("rows"))
    job = payload.get("job") or {}
    mapping = payload.get("mapping_json") or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"
    total_cols = 8
    merge_write(ws, 1, 1, 1, total_cols, mapping.get("title") or "견 적 서", bold=True, size=20, fill=TITLE_FILL)
    write_cell(ws, 3, 1, "견적일자", bold=True, fill=HEADER_FILL)
    write_cell(ws, 3, 2, today_text())
    write_cell(ws, 3, 3, "수신", bold=True, fill=HEADER_FILL)
    merge_write(ws, 3, 4, 3, 5, rows[0].get("recipient", "") if rows else "")
    write_cell(ws, 3, 6, "작성자", bold=True, fill=HEADER_FILL)
    merge_write(ws, 3, 7, 3, 8, payload.get("author_name") or "")
    write_cell(ws, 4, 1, "공급자", bold=True, fill=HEADER_FILL)
    merge_write(ws, 4, 2, 4, 4, rows[0].get("vendor_name", "") if rows else "")
    write_cell(ws, 4, 5, "견적명", bold=True, fill=HEADER_FILL)
    merge_write(ws, 4, 6, 4, 8, rows[0].get("document_title") or job.get("title") or "")
    headers = ["NO", "품명", "규격", "수량", "단위", "단가", "금액", "비고"]
    for idx, header in enumerate(headers, start=1):
        write_cell(ws, 6, idx, header, bold=True, fill=HEADER_FILL)
    total = 0
    for row_idx, row in enumerate(rows, start=7):
        values = [row_idx - 6, get_row_value(row, "item_name", row_idx-7), get_row_value(row, "spec"), get_row_value(row, "quantity"), get_row_value(row, "unit"), get_row_value(row, "unit_price"), get_row_value(row, "amount"), get_row_value(row, "remark")]
        total += to_number(values[6])
        for col_idx, value in enumerate(values, start=1):
            write_cell(ws, row_idx, col_idx, value, align="left" if col_idx in (2, 8) else "center")
            if col_idx in (6, 7):
                ws.cell(row_idx, col_idx).number_format = "#,##0"
    total_row = 7 + len(rows)
    merge_write(ws, total_row, 1, total_row, 6, "합계", bold=True, fill=LIGHT_FILL)
    write_cell(ws, total_row, 7, int(total) if total else "", bold=True, fill=LIGHT_FILL)
    ws.cell(total_row, 7).number_format = "#,##0"
    write_cell(ws, total_row, 8, "", fill=LIGHT_FILL)
    merge_write(ws, total_row + 2, 1, total_row + 2, 2, "특기사항", bold=True, fill=HEADER_FILL)
    merge_write(ws, total_row + 2, 3, total_row + 3, 8, rows[0].get("special_note", "") if rows else "", fill=TITLE_FILL)
    set_widths(ws, total_cols, 14)
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["H"].width = 24
    return wb, {"template_kind": "ESTIMATE_FORM"}


def create_price_table_workbook(payload: Dict[str, Any]) -> Tuple[Workbook, Dict[str, Any]]:
    rows = as_list(payload.get("rows"))
    columns = [
        {"key": "row_no", "label": "NO"}, {"key": "construction_code", "label": "공종코드"},
        {"key": "item_name", "label": "공종명/품명"}, {"key": "spec", "label": "규격"},
        {"key": "unit", "label": "단위"}, {"key": "quantity", "label": "수량"},
        {"key": "unit_price", "label": "기준단가"}, {"key": "amount", "label": "금액"}, {"key": "remark", "label": "비고"},
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "단가표"
    merge_write(ws, 1, 1, 1, len(columns), "표준 단가표", bold=True, size=16, fill=TITLE_FILL)
    write_cell(ws, 3, 1, "기준일", bold=True, fill=HEADER_FILL)
    write_cell(ws, 3, 2, today_text())
    write_cell(ws, 3, 3, "적용범위", bold=True, fill=HEADER_FILL)
    merge_write(ws, 3, 4, 3, len(columns), "공사/자재/장비 단가 관리")
    write_table(ws, columns, rows, 5)
    return wb, {"template_kind": "UNIT_PRICE_TABLE", "columns": columns}


def resolve_binding_value(payload: Dict[str, Any], row: Dict[str, Any], binding_key: str, default: str = "") -> Any:
    key = str(binding_key or "").strip()
    if not key:
        return default
    if key in ("today", "document_date"):
        return row.get(key) or today_text()
    if key in row and row.get(key) not in (None, ""):
        return row.get(key)
    analysis = (payload.get("job") or {}).get("analysis") or {}
    camel = re.sub(r"_([a-z])", lambda m: m.group(1).upper(), key)
    for candidate in (key, camel):
        if isinstance(analysis, dict) and analysis.get(candidate) not in (None, ""):
            return analysis.get(candidate)
    if key in ("requester_name", "writer_name", "created_by"):
        return payload.get("author_name") or default
    if key in ("summary", "purpose", "content"):
        return analysis.get("summary") or analysis.get("purpose") or (payload.get("job") or {}).get("userRequest") or default
    if key in ("review_opinion", "review", "issue_summary"):
        return analysis.get("reviewSummary") or analysis.get("review_summary") or analysis.get("summary") or default
    if key == "action_plan":
        return analysis.get("actionPlan") or analysis.get("nextSteps") or default
    return default


def create_custom_document_workbook(payload: Dict[str, Any]) -> Tuple[Workbook, Dict[str, Any]]:
    """Gemini가 만든 sections/headerPairs를 그대로 렌더링하는 범용 회사 문서형 엑셀."""
    rows = as_list(payload.get("rows"))
    first = rows[0] if rows else {}
    mapping = payload.get("mapping_json") or {}
    sections = as_list(mapping.get("sections"))
    if not sections:
        sections = [
            {"title": "1. 작성 목적", "bindingKey": "purpose", "height": 3},
            {"title": "2. 주요 내용", "bindingKey": "summary", "height": 4},
            {"title": "3. 검토 의견", "bindingKey": "review_opinion", "height": 4},
            {"title": "4. 후속 조치", "bindingKey": "action_plan", "height": 3},
        ]
    header_pairs = as_list(mapping.get("headerPairs")) or [
        {"label": "작성일", "bindingKey": "document_date"},
        {"label": "작성자", "bindingKey": "requester_name"},
        {"label": "공사명", "bindingKey": "project_name"},
        {"label": "현장명", "bindingKey": "site_name"},
    ]
    approvals = [str(x) for x in as_list(mapping.get("approvalLines")) if str(x).strip()] or ["담당", "검토", "승인"]
    total_cols = 8
    wb = Workbook()
    ws = wb.active
    ws.title = str(mapping.get("sheetName") or mapping.get("sheet") or "AI문서양식")[:31]

    title = str(mapping.get("title") or first.get("document_title") or first.get("report_title") or "AI 생성 문서 양식")
    merge_write(ws, 1, 1, 1, total_cols, title, bold=True, size=18, fill=TITLE_FILL)

    # 우측 결재란
    start_approval_col = max(5, total_cols - len(approvals) + 1)
    for idx, label in enumerate(approvals[:4], start=start_approval_col):
        write_cell(ws, 2, idx, label, bold=True, fill=HEADER_FILL)
        write_cell(ws, 3, idx, "")
        ws.row_dimensions[3].height = 32

    r = 3
    c = 1
    for pair in header_pairs[:8]:
        label = str(pair.get("label") or "").strip()
        binding = str(pair.get("bindingKey") or pair.get("fieldKey") or "").strip()
        if not label or not binding:
            continue
        write_cell(ws, r, c, label, bold=True, fill=HEADER_FILL)
        merge_write(ws, r, c + 1, r, min(c + 2, total_cols), resolve_binding_value(payload, first, binding, ""))
        c += 4
        if c > total_cols:
            r += 1
            c = 1

    r += 2
    for idx, section in enumerate(sections[:12], start=1):
        section_title = str(section.get("title") or section.get("label") or f"{idx}. 문서 내용")
        binding = str(section.get("bindingKey") or section.get("fieldKey") or section.get("key") or "summary")
        height = section.get("height", 3)
        try:
            height = int(height)
        except Exception:
            height = 3
        height = max(1, min(8, height))
        merge_write(ws, r, 1, r, total_cols, section_title, bold=True, fill=HEADER_FILL)
        value = resolve_binding_value(payload, first, binding, "")
        merge_write(ws, r + 1, 1, r + height, total_cols, value or "")
        for rr in range(r + 1, r + height + 1):
            ws.row_dimensions[rr].height = 24
        r += height + 2

    # 표 컬럼이 같이 필요한 문서 양식이면 하단에 첨부 내역표를 붙인다.
    base_columns = as_list(mapping.get("baseColumns"))
    if base_columns and rows:
        merge_write(ws, r, 1, r, total_cols, "첨부 내역", bold=True, fill=HEADER_FILL)
        norm_cols = []
        for item in base_columns[:total_cols]:
            key = normalize_key(item.get("fieldKey") or item.get("key"))
            if key:
                norm_cols.append({"key": key, "label": label_for(key, item.get("label"))})
        if norm_cols:
            write_table(ws, norm_cols, rows, r + 1)

    set_widths(ws, total_cols, 16)
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    return wb, {"template_kind": "CUSTOM_DOCUMENT_FORM", "section_count": len(sections)}

def create_report_workbook(payload: Dict[str, Any], doc_type: str) -> Tuple[Workbook, Dict[str, Any]]:
    rows = as_list(payload.get("rows"))
    first = rows[0] if rows else {}
    wb = Workbook()
    ws = wb.active
    ws.title = DOCUMENT_TYPE_LABELS.get(doc_type, "보고서")[:31]

    if doc_type == "MEETING_MINUTES":
        first, action_rows = build_meeting_rows(payload, rows)
        total_cols = 6
        merge_write(ws, 1, 1, 1, total_cols, first.get("meeting_title") or "회 의 록", bold=True, size=18, fill=TITLE_FILL)
        write_cell(ws, 3, 1, "회의일시", bold=True, fill=HEADER_FILL)
        write_cell(ws, 3, 2, first.get("meeting_date") or today_text())
        write_cell(ws, 3, 3, "장소", bold=True, fill=HEADER_FILL)
        write_cell(ws, 3, 4, first.get("meeting_place", ""))
        write_cell(ws, 3, 5, "작성자", bold=True, fill=HEADER_FILL)
        write_cell(ws, 3, 6, payload.get("author_name") or "")
        write_cell(ws, 4, 1, "참석자", bold=True, fill=HEADER_FILL)
        merge_write(ws, 4, 2, 4, total_cols, first.get("attendees", ""))
        sections = [("1. 회의 안건", first.get("agenda") or get_row_value(first, "item_name")), ("2. 주요 논의 내용", first.get("discussion") or first.get("content") or ""), ("3. 결정 사항", first.get("decision") or ""), ("4. 비고", first.get("remark") or "")]
        r = 6
        for title, body in sections:
            merge_write(ws, r, 1, r, total_cols, title, bold=True, fill=HEADER_FILL)
            merge_write(ws, r + 1, 1, r + 3, total_cols, body or "")
            r += 5
        headers = ["NO", "조치내용", "담당자", "기한", "상태", "비고"]
        for c, h in enumerate(headers, start=1):
            write_cell(ws, r, c, h, bold=True, fill=HEADER_FILL)
        for idx, row in enumerate(action_rows, start=1):
            vals = [idx, row.get("action_item") or "", row.get("owner") or row.get("manager") or "확인 필요", row.get("due_date") or "미정", row.get("status") or "확인 필요", row.get("remark") or ""]
            for c, v in enumerate(vals, start=1):
                write_cell(ws, r + idx, c, v, align="left" if c in (2, 6) else "center")
        set_widths(ws, total_cols, 16)
        ws.column_dimensions["B"].width = 34
        return wb, {"template_kind": doc_type}

    if doc_type == "OFFICIAL_LETTER":
        first = build_official_first_row(payload, rows)
        total_cols = 6
        merge_write(ws, 1, 1, 1, total_cols, first.get("letter_title") or "공 문", bold=True, size=20, fill=TITLE_FILL)
        write_cell(ws, 3, 1, "문서번호", bold=True, fill=HEADER_FILL)
        merge_write(ws, 3, 2, 3, 3, first.get("document_no", ""))
        write_cell(ws, 3, 4, "시행일자", bold=True, fill=HEADER_FILL)
        merge_write(ws, 3, 5, 3, 6, today_text())
        write_cell(ws, 4, 1, "수신", bold=True, fill=HEADER_FILL)
        merge_write(ws, 4, 2, 4, 6, first.get("recipient", ""))
        write_cell(ws, 5, 1, "참조", bold=True, fill=HEADER_FILL)
        merge_write(ws, 5, 2, 5, 6, first.get("reference", ""))
        write_cell(ws, 6, 1, "제목", bold=True, fill=HEADER_FILL)
        merge_write(ws, 6, 2, 6, 6, first.get("document_title") or first.get("title") or "")
        write_cell(ws, 8, 1, "본문", bold=True, fill=HEADER_FILL)
        merge_write(ws, 8, 2, 14, 6, first.get("body") or first.get("content") or first.get("summary") or "")
        write_cell(ws, 16, 1, "붙임", bold=True, fill=HEADER_FILL)
        merge_write(ws, 16, 2, 16, 6, first.get("attachment_note", ""))
        merge_write(ws, 19, 1, 19, 6, first.get("sender") or "공사팀", bold=True, size=14, fill=TITLE_FILL)
        set_widths(ws, total_cols, 16)
        ws.column_dimensions["B"].width = 22
        return wb, {"template_kind": doc_type}

    # BUSINESS REPORT / GENERAL REPORT
    first = build_report_first_row(payload, rows)
    total_cols = 8
    title = first.get("report_title") or first.get("document_title") or "업무 보고서"
    merge_write(ws, 1, 1, 1, total_cols, title, bold=True, size=18, fill=TITLE_FILL)
    write_cell(ws, 3, 1, "작성일", bold=True, fill=HEADER_FILL)
    write_cell(ws, 3, 2, today_text())
    write_cell(ws, 3, 3, "작성자", bold=True, fill=HEADER_FILL)
    write_cell(ws, 3, 4, payload.get("author_name") or "")
    write_cell(ws, 3, 5, "문서구분", bold=True, fill=HEADER_FILL)
    merge_write(ws, 3, 6, 3, 8, "보고서")
    sections = [
        ("1. 보고 목적", first.get("report_purpose") or first.get("purpose") or ""),
        ("2. 주요 검토 내용", first.get("summary") or first.get("content") or ""),
        ("3. 주요 이슈 및 확인사항", first.get("issue_summary") or first.get("review_result") or first.get("review_opinion") or ""),
        ("4. 후속 조치 및 관리계획", first.get("action_plan") or ""),
    ]
    r = 5
    for title, body in sections:
        merge_write(ws, r, 1, r, total_cols, title, bold=True, fill=HEADER_FILL)
        merge_write(ws, r + 1, 1, r + 3, total_cols, body or "")
        r += 5
    merge_write(ws, r, 1, r, total_cols, "5. 참고 사항", bold=True, fill=HEADER_FILL)
    merge_write(ws, r + 1, 1, r + 3, total_cols, first.get("footer_note") or "")
    set_widths(ws, total_cols, 16)
    ws.column_dimensions["B"].width = 26
    return wb, {"template_kind": doc_type}

def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def resolve_template_path(template: Dict[str, Any]) -> Optional[Path]:
    raw = template.get("file_path") or template.get("filePath") or ""
    if not raw:
        return None
    candidates = [Path(raw), Path.cwd() / raw, BASE_DIR / raw, BASE_DIR.parent / raw]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def cell_col(cell_address: str) -> Optional[int]:
    m = re.match(r"^([A-Z]+)", str(cell_address or ""), re.I)
    return column_index_from_string(m.group(1).upper()) if m else None


def cell_row(cell_address: str) -> Optional[int]:
    m = re.search(r"(\d+)$", str(cell_address or ""))
    return int(m.group(1)) if m else None


def create_mapped_template_workbook(payload: Dict[str, Any]) -> Optional[Tuple[Workbook, Dict[str, Any]]]:
    template = payload.get("template") or {}
    mapping_json = payload.get("mapping_json") or {}
    mappings = as_list(payload.get("mappings") or mapping_json.get("mappings"))
    template_path = resolve_template_path(template)
    if not template_path or not mappings:
        return None
    try:
        wb = load_workbook(template_path)
    except Exception:
        return None
    ws = wb[mappings[0].get("sheetName")] if mappings and mappings[0].get("sheetName") in wb.sheetnames else wb.active
    rows = as_list(payload.get("rows"))
    columns = as_list(payload.get("columns"))
    job = payload.get("job") or {}
    table_json = {}
    if isinstance(job.get("tables"), list) and job["tables"]:
        table_json = job["tables"][0].get("tableJson") or job["tables"][0].get("table_json") or {}
    vendors = infer_vendors(columns, rows, table_json)

    for m in mappings:
        field_key = m.get("fieldKey") or m.get("field_key")
        mapping_type = str(m.get("mappingType") or m.get("mapping_type") or "").upper()
        if mapping_type == "SINGLE_CELL":
            addr = m.get("cellAddress") or str(m.get("mergedRange") or "").split(":")[0]
            if not addr:
                continue
            val = today_text() if field_key == "document_date" else (payload.get("author_name") if field_key in ("requester_name", "writer_name", "created_by") else get_row_value(rows[0] if rows else {}, field_key))
            ws[addr] = val
        elif mapping_type in ("REPEAT_ROW", "REPEAT_COLUMN"):
            col_letter = m.get("columnLetter") or re.sub(r"\d", "", str(m.get("cellAddress") or ""))
            if not col_letter:
                continue
            col = column_index_from_string(str(col_letter).upper())
            start = int(m.get("startRow") or cell_row(m.get("cellAddress")) or 7)
            max_rows = int(m.get("maxRows") or len(rows) or 1)
            for idx in range(max_rows):
                val = get_row_value(rows[idx], field_key, idx) if idx < len(rows) else None
                ws.cell(start + idx, col).value = val
        elif mapping_type == "COMPANY_GROUP_COLUMN":
            start = int(m.get("startRow") or 7)
            field = field_key
            letters = as_list(m.get("columnLetters"))
            group_width = int(m.get("groupWidth") or 4)
            if field == "target_name":
                ranges = as_list(m.get("groupRanges"))
                for idx, vendor in enumerate(vendors):
                    if idx < len(ranges) and ranges[idx]:
                        addr = str(ranges[idx]).split(":")[0]
                        ws[addr] = vendor.get("name")
                    elif letters:
                        ws[f"{letters[0]}{start}"] = vendor.get("name")
                continue
            for vidx, vendor in enumerate(vendors):
                if letters and vidx < len(letters):
                    col = column_index_from_string(letters[vidx])
                elif letters:
                    col = column_index_from_string(letters[0]) + vidx * group_width
                else:
                    col = 6 + vidx * group_width
                for ridx, row in enumerate(rows):
                    ws.cell(start + ridx, col).value = get_vendor_value(row, vendor, field)
    return wb, {"template_kind": "MAPPED_COMPANY_TEMPLATE", "vendor_count": len(vendors)}


def create_design_workbook(payload: Dict[str, Any]) -> Tuple[Workbook, Dict[str, Any]]:
    mapping = payload.get("mapping_json") or {}
    layout = str(mapping.get("layout") or mapping.get("layoutType") or "").upper()
    doc_type = detect_document_type(payload)

    # 디자인 후보를 선택한 경우 문서 유형보다 layout을 우선한다.
    # 이렇게 해야 '업체 비교형/요약 우선형/기본 표형'이 이름만 다른 동일 양식으로 생성되지 않는다.
    if "DYNAMIC_VENDOR" in layout or "VENDOR_COMPARE" in layout or layout == "AI_GENERATED_DYNAMIC_VENDOR_TABLE":
        return create_estimate_comparison_workbook(payload)
    if "ESTIMATE" in layout:
        return create_estimate_form_workbook(payload)
    if "PRICE" in layout or "UNIT_PRICE" in layout:
        return create_price_table_workbook(payload)
    if "CUSTOM_DOCUMENT_FORM" in layout or "DOCUMENT_FORM" in layout:
        return create_custom_document_workbook(payload)
    if "OFFICIAL" in layout:
        return create_report_workbook(payload, "OFFICIAL_LETTER")
    if "MEETING" in layout:
        return create_report_workbook(payload, "MEETING_MINUTES")
    if any(token in layout for token in ("SECTION", "SUMMARY", "APPROVAL", "HEADER_TABLE", "REPORT")):
        return create_report_workbook(payload, "REPORT")
    if "TABLE_ONLY" in layout:
        return create_free_form_workbook(payload)

    if doc_type == "ESTIMATE_COMPARISON":
        return create_estimate_comparison_workbook(payload)
    if doc_type in ("REPORT", "MEETING_MINUTES", "OFFICIAL_LETTER"):
        return create_report_workbook(payload, doc_type)
    if doc_type == "UNIT_PRICE_TABLE":
        return create_price_table_workbook(payload)
    return create_free_form_workbook(payload)


def save_workbook(wb: Workbook, file_name: Optional[str]) -> Dict[str, Any]:
    ensure_dir(RESULT_DIR)
    safe = safe_filename(file_name, f"excel_result_{now_stamp()}.xlsx")
    out_path = RESULT_DIR / f"{now_stamp()}_{safe}"
    wb.save(out_path)
    return {"file_name": safe, "file_path": str(out_path)}


def generate_excel(payload: Dict[str, Any]) -> Dict[str, Any]:
    mapped = None
    mapping_json = payload.get("mapping_json") or {}
    if payload.get("output_mode") == "COMPANY_TEMPLATE" and not mapping_json.get("aiGenerated"):
        mapped = create_mapped_template_workbook(payload)
    if mapped:
        wb, meta = mapped
    else:
        wb, meta = create_design_workbook(payload)
    result = save_workbook(wb, payload.get("file_name"))
    result.update({
        "template_kind": meta.get("template_kind"),
        "vendor_count": meta.get("vendor_count", 0),
        "engine": "openpyxl",
    })
    return result


def make_template_skeleton(payload: Dict[str, Any]) -> Dict[str, Any]:
    design = payload.get("design") or {}
    skeleton_payload = {
        "rows": [],
        "columns": design.get("baseColumns") or [],
        "mapping_json": design,
        "template": {"template_name": design.get("templateName"), "template_type": design.get("templateType")},
        "job": {"user_request": design.get("reason") or design.get("title") or ""},
        "output_mode": "COMPANY_TEMPLATE",
        "file_name": payload.get("file_name") or f"{design.get('templateName') or 'AI_TEMPLATE'}.xlsx",
    }
    wb, meta = create_design_workbook(skeleton_payload)
    ensure_dir(TEMPLATE_DIR)
    safe = safe_filename(payload.get("file_name"), f"ai_template_{now_stamp()}.xlsx")
    out_path = TEMPLATE_DIR / f"{now_stamp()}_{safe}"
    wb.save(out_path)
    return {"file_name": safe, "file_path": str(out_path), "template_kind": meta.get("template_kind"), "engine": "openpyxl"}


def build_design_candidates(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    이전 layout registry 기반 후보 생성 기능은 사용하지 않는다.
    Gemini가 실제 양식 JSON을 생성하고, 서버가 sanitize 후 openpyxl 렌더링한다.
    """
    analysis = payload.get("analysis") or {}
    doc_type = str(analysis.get("documentType") or analysis.get("document_type") or "업무 문서")
    return {"document_type": doc_type, "design_candidates": []}


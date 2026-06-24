from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ── 경로 상수 ────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parents[3]
RESULT_DIR = BASE_DIR / "storage" / "results"
TEMPLATE_DIR = BASE_DIR / "storage" / "templates" / "ai_generated"

# ── 스타일 상수 ──────────────────────────────────────────────────────────────
THIN = Side(style="thin", color="B0BEC5")
MEDIUM = Side(style="medium", color="607D8B")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
OUTER_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")   # 진한 남색
HEADER_FILL2 = PatternFill("solid", fgColor="34495E")  # 업체 헤더 구분색
LIGHT_FILL = PatternFill("solid", fgColor="EBF5FB")    # 연한 파랑
TITLE_FILL = PatternFill("solid", fgColor="1A252F")    # 제목 배경 (진한 네이비)
GREEN_FILL = PatternFill("solid", fgColor="D5F5E3")
AMBER_FILL = PatternFill("solid", fgColor="FEF9E7")
ALT_ROW_FILL = PatternFill("solid", fgColor="F8FBFF")  # 짝수행 연한 줄무늬

DOCUMENT_TYPE_LABELS = {
    "ESTIMATE": "견적서",
    "ESTIMATE_COMPARISON": "비교 견적서",
    "MULTI_VENDOR_PRICE_COMPARISON": "비교 견적서",
    "UNIT_PRICE_TABLE": "단가표",
    "STANDARD_MARKET_PRICE_TABLE": "단가표",
    "REPORT": "보고서",
    "BUSINESS_REPORT": "보고서",
    "MEETING_MINUTES": "회의록",
    "OFFICIAL_LETTER": "공문",
    "NORMAL_TABLE": "문서 정리표",
    "CUSTOM_DOCUMENT_FORM": "문서 양식",
}

TYPE_KEYWORDS = {
    "ESTIMATE_COMPARISON": ["비교", "견적", "업체", "회사", "단가비교", "가격비교"],
    "UNIT_PRICE_TABLE": ["단가표", "단가", "표준시장", "품셈", "가격표"],
    "REPORT": ["보고서", "보고", "현황", "요약", "검토"],
    "MEETING_MINUTES": ["회의록", "회의", "안건", "참석", "결정사항"],
    "OFFICIAL_LETTER": ["공문", "수신", "참조", "제목", "시행"],
}

BASE_LABELS = {
    "row_no": "NO",
    "no": "NO",
    "project_name": "공사명",
    "site_name": "현장명",
    "document_title": "제목",
    "document_date": "작성일",
    "requester_name": "작성자",
    "writer_name": "작성자",
    "vendor_name": "업체명",
    "company_name": "회사명",
    "construction_code": "공종코드",
    "item_name": "품명",
    "product_name": "품명",
    "work_item_name": "공종명칭",
    "spec": "규격",
    "standard": "규격",
    "quantity": "수량",
    "unit": "단위",
    "unit_price": "단가",
    "standard_unit_price": "표준단가",
    "vendor_unit_price": "업체단가",
    "amount": "금액",
    "total_amount": "합계",
    "lowest_target": "최저 업체",
    "lowest_vendor": "최저 업체",
    "calculated_unit_price": "최저 단가",
    "lowest_unit_price": "최저 단가",
    "remark": "비고",
    "note": "비고",
    "meeting_date": "회의일자",
    "attendees": "참석자",
    "agenda": "안건",
    "decision": "결정사항",
    "action_item": "조치사항",
    "recipient": "수신",
    "reference": "참조",
    "sender": "발신",
    "body": "내용",
    "summary": "요약",
    "opinion": "검토의견",
}


def request_output_intent(text: str = "") -> str:
    raw = str(text or "")
    wants_report = any(token in raw for token in ["보고서 형식", "보고서", "업무보고서", "검토보고서", "서술형", "문장형", "보고용"])
    wants_table = any(token in raw for token in ["표로", "표 형태", "표 형식", "비교표", "단가표", "조사표", "테이블", "그리드", "엑셀 표"])
    if wants_table and not wants_report:
        return "TABLE"
    if wants_table and re.search(r"표로|비교표|단가표|조사표|테이블|그리드|엑셀\s*표", raw):
        return "TABLE"
    if wants_report:
        return "REPORT"
    return "AUTO"


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def safe_filename(name: Optional[str], fallback: str = "document_result.xlsx") -> str:
    raw = str(name or fallback).strip() or fallback
    raw = re.sub(r'[\\/:*?"<>|]+', "_", raw)
    if not raw.lower().endswith(".xlsx"):
        raw += ".xlsx"
    return raw


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def today_text() -> str:
    return datetime.now().strftime("%Y.%m.%d")


def as_list(value: Any) -> List[Any]:
    return value if isinstance(value, list) else []


def normalize_key(value: Any) -> str:
    return str(value or "").strip()


def label_for(key: str, fallback: Optional[str] = None) -> str:
    return str(fallback or BASE_LABELS.get(key) or key)


def to_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    raw = re.sub(r"[^0-9.\-]", "", str(value or ""))
    try:
        return float(raw) if raw not in ("", "-", ".") else 0.0
    except Exception:
        return 0.0


_DARK_FILLS = {HEADER_FILL, HEADER_FILL2, TITLE_FILL}


def _font_color_for(fill: Optional[PatternFill]) -> str:
    return "FFFFFF" if fill in _DARK_FILLS else "1A252F"


def write_cell(ws: Worksheet, row: int, col: int, value: Any, *, bold: bool = False, fill: Optional[PatternFill] = None,
               align: str = "center", border: bool = True, wrap: bool = True) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    color = _font_color_for(fill)
    cell.font = Font(bold=bold, size=11, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    if border:
        cell.border = BORDER


def merge_write(ws: Worksheet, start_row: int, start_col: int, end_row: int, end_col: int, value: Any,
                *, bold: bool = False, fill: Optional[PatternFill] = None, size: int = 11) -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = ws.cell(start_row, start_col, value)
    color = _font_color_for(fill)
    cell.font = Font(bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).border = BORDER


def set_widths(ws: Worksheet, count: int, base: int = 14) -> None:
    for c in range(1, count + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(base, min(28, base + (c % 3)))


def auto_fit_columns(ws: Worksheet, min_width: int = 10, max_width: int = 40) -> None:
    """모든 셀 내용을 기준으로 컬럼 너비를 자동 조정"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                text = str(cell.value)
                # 한글은 2자 너비로 계산
                cell_len = sum(2 if ord(c) > 127 else 1 for c in text)
                max_len = max(max_len, cell_len)
        adjusted = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = adjusted


def apply_print_settings(ws: Worksheet, total_cols: int, total_rows: int, header_row: int = 1) -> None:
    """인쇄 영역, 페이지 설정, 여백, 헤더고정 적용"""
    from openpyxl.worksheet.page import PageMargins
    # 인쇄 영역 설정
    last_col = get_column_letter(total_cols)
    ws.print_area = f"A1:{last_col}{total_rows}"
    # 페이지 방향 (컬럼 수가 많으면 가로)
    ws.page_setup.orientation = 'landscape' if total_cols > 8 else 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = 9  # A4
    # 인쇄 여백
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3)
    # 헤더 행 반복 인쇄
    ws.print_title_rows = f"1:{header_row}"


def get_row_value(row: Dict[str, Any], key: str, index: int = 0) -> Any:
    aliases = {
        "row_no": ["row_no", "no", "순번"],
        "item_name": ["item_name", "product_name", "work_item_name", "품명", "품목", "공종명칭", "항목"],
        "spec": ["spec", "standard", "규격", "사양"],
        "quantity": ["quantity", "qty", "request_quantity", "requested_quantity", "수량"],
        "unit": ["unit", "단위"],
        "unit_price": ["unit_price", "vendor_unit_price", "price", "단가"],
        "amount": ["amount", "total_amount", "금액"],
        "remark": ["remark", "note", "memo", "비고"],
    }
    if key in ("row_no", "no"):
        return row.get("row_no") or row.get("no") or index + 1
    for alias in aliases.get(key, [key]):
        if alias in row and row.get(alias) not in (None, ""):
            return row.get(alias)
    val = row.get(key)
    if val not in (None, ""):
        return val
    return ""


def first_text(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        if isinstance(value, list):
            parts = []
            for item in value:
                if isinstance(item, dict):
                    parts.append(str(item.get("value") or item.get("content") or item.get("text") or ""))
                else:
                    parts.append(str(item or ""))
            text = "\n".join(part.strip() for part in parts if part and str(part).strip()).strip()
            if text:
                return text
            continue
        if isinstance(value, dict):
            text = str(value.get("value") or value.get("content") or value.get("text") or value.get("summary") or "").strip()
            if text:
                return text
            continue
        text = str(value or "").strip()
        if text:
            return text
    return ""


def compact_status(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_json_maybe(value: Any) -> Any:
    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return {}
    return value if isinstance(value, dict) else {}


def get_payload_analysis(payload: Dict[str, Any]) -> Dict[str, Any]:
    job = parse_json_maybe(payload.get("job"))
    analysis = parse_json_maybe(job.get("analysis"))
    raw = parse_json_maybe(analysis.get("raw"))
    merged = {**raw, **analysis}
    return merged


def get_payload_drafts(payload: Dict[str, Any]) -> Dict[str, Any]:
    analysis = get_payload_analysis(payload)
    drafts = analysis.get("drafts") if isinstance(analysis.get("drafts"), dict) else {}
    return drafts


def ignore_plain_status(value: Any) -> bool:
    text = compact_status(value)
    return not text or text in {"정상", "완료", "없음", "해당없음", "확인"} or len(text) <= 2


def looks_like_user_prompt_or_meta(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    prompt_tokens = [
        "정리해줘", "작성해줘", "만들어줘", "써줘", "출력해줘", "보여줘",
        "보고서 형식", "보고서 형태", "보고서로", "요약과 검토내용", "상세하게", "풍부하게",
    ]
    meta_pattern = r"첨부\s*파일|총\s*페이지|확인된\s*총\s*페이지|표\s*후보|PyMuPDF|pdfplumber|PP-Structure|PaddleOCR|OCR|LLM|ai-server|저장\s*위치|산출\s*방식|요청\s*내용은|백그라운드|엑셀\s*미리보기"
    return any(token in text for token in prompt_tokens) or bool(re.search(meta_pattern, text, re.I))


def document_only_text(value: Any) -> str:
    text = str(value or "").strip()
    if not text or looks_like_user_prompt_or_meta(text):
        return ""
    text = re.sub(r"\[page\s*\d+\s*/\s*\d+(?:\s*OCR)?\]", "", text, flags=re.I)
    text = re.sub(r"텍스트\s*전용\s*샘플\s*\d*", "", text).strip()
    lines = [line.strip() for line in re.split(r"\r?\n", text) if line.strip()]
    clean_lines = [line for line in lines if not looks_like_user_prompt_or_meta(line)]
    return "\n".join(clean_lines)


def _strip_page_markers(text: str) -> str:
    """PDF 페이지 마커와 샘플 문서 표시를 제거한다."""
    text = re.sub(r"\[page\s*\d+\s*/\s*\d+(?:\s*OCR)?\]", "", text, flags=re.I)
    text = re.sub(r"텍스트\s*전용\s*샘플\s*\d*", "", text, flags=re.I)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


# ── detect / normalize helpers ───────────────────────────────────────────────
def detect_document_type(payload: Dict[str, Any]) -> str:
    mapping = payload.get("mapping_json") or {}
    template = payload.get("template") or {}
    job = payload.get("job") or {}
    analysis = job.get("analysis") or {}
    table_type = str((payload.get("job") or {}).get("tables", [{}])[0].get("tableType", "") if isinstance((payload.get("job") or {}).get("tables"), list) else "")
    text = " ".join(str(x or "") for x in [
        payload.get("output_mode"), mapping.get("template_type"), mapping.get("templateType"), mapping.get("layout"),
        template.get("template_type"), template.get("templateType"), template.get("template_name"), template.get("templateName"),
        job.get("userRequest"), job.get("user_request"), analysis.get("documentType"), analysis.get("recommendedTableType"), table_type,
    ]).upper()
    korean_text = " ".join(str(x or "") for x in [template.get("template_name"), template.get("templateName"), job.get("userRequest"), job.get("user_request")])
    if "MULTI_VENDOR" in text or "COMPARISON" in text or any(k in korean_text for k in TYPE_KEYWORDS["ESTIMATE_COMPARISON"]):
        return "ESTIMATE_COMPARISON"
    if "UNIT_PRICE" in text or "STANDARD_MARKET" in text or any(k in korean_text for k in TYPE_KEYWORDS["UNIT_PRICE_TABLE"]):
        return "UNIT_PRICE_TABLE"
    if "CUSTOM_DOCUMENT_FORM" in text or "DOCUMENT_FORM" in text:
        return "CUSTOM_DOCUMENT_FORM"
    if "MEETING" in text or any(k in korean_text for k in TYPE_KEYWORDS["MEETING_MINUTES"]):
        return "MEETING_MINUTES"
    if "OFFICIAL" in text or any(k in korean_text for k in TYPE_KEYWORDS["OFFICIAL_LETTER"]):
        return "OFFICIAL_LETTER"
    if "REPORT" in text or any(k in korean_text for k in TYPE_KEYWORDS["REPORT"]):
        return "REPORT"
    return "NORMAL_TABLE"


def normalize_columns(columns: List[Dict[str, Any]], rows: List[Dict[str, Any]], doc_type: str, mapping_json: Optional[Dict[str, Any]] = None) -> List[Dict[str, str]]:
    mapping_json = mapping_json or {}
    if mapping_json.get("baseColumns"):
        out = []
        for item in as_list(mapping_json.get("baseColumns")):
            key = normalize_key(item.get("fieldKey") or item.get("key"))
            label = item.get("label") or item.get("fieldLabel") or label_for(key)
            if key:
                out.append({"key": key, "label": label})
        if not any(c["key"] in ("row_no", "no") for c in out):
            out.insert(0, {"key": "row_no", "label": "NO"})
        return out

    seen: set = set()
    out: List[Dict[str, str]] = []
    preferred = {
        "ESTIMATE_COMPARISON": ["row_no", "item_name", "spec", "quantity", "unit"],
        "UNIT_PRICE_TABLE": ["row_no", "construction_code", "item_name", "spec", "unit", "unit_price", "amount", "remark"],
        "REPORT": ["row_no", "section", "summary", "item_name", "content", "remark"],
        "MEETING_MINUTES": ["row_no", "agenda", "decision", "action_item", "owner", "due_date", "remark"],
        "OFFICIAL_LETTER": ["row_no", "recipient", "reference", "document_title", "body", "remark"],
        "NORMAL_TABLE": ["row_no", "item_name", "spec", "quantity", "unit", "unit_price", "amount", "remark"],
    }.get(doc_type, [])

    def add(key: str, label: Optional[str] = None) -> None:
        if not key or key in seen:
            return
        seen.add(key)
        out.append({"key": key, "label": label_for(key, label)})

    for key in preferred:
        add(key)
    for col in columns or []:
        key = normalize_key(col.get("key") or col.get("fieldKey"))
        if re.match(r"^(vendor|company|target)_?\d+_", key, re.I):
            continue
        add(key, col.get("label"))
    for row in rows[:10] if rows else []:
        for key, val in row.items():
            if val not in (None, "") and not re.match(r"^(vendor|company|target)_?\d+_", key, re.I):
                add(key)
    return out[:18]


def write_title_area(ws: Worksheet, title: str, total_cols: int, subtitle: str = "") -> None:
    ws.row_dimensions[1].height = 34
    merge_write(ws, 1, 1, 1, total_cols, title, bold=True, size=16, fill=TITLE_FILL)
    if subtitle:
        ws.row_dimensions[2].height = 16
        merge_write(ws, 2, 1, 2, total_cols, subtitle, bold=False, size=9, fill=LIGHT_FILL)
    write_cell(ws, 3, 1, "작성일", bold=True, fill=HEADER_FILL)
    write_cell(ws, 3, 2, today_text(), fill=LIGHT_FILL)
    write_cell(ws, 3, 3, "작성자", bold=True, fill=HEADER_FILL)
    write_cell(ws, 3, 4, "", fill=LIGHT_FILL)


def write_table(ws: Worksheet, columns: List[Dict[str, str]], rows: List[Dict[str, Any]], start_row: int = 5) -> None:
    for idx, col in enumerate(columns, start=1):
        write_cell(ws, start_row, idx, col.get("label") or col.get("key"), bold=True, fill=HEADER_FILL)
    for r_idx, row in enumerate(rows, start=start_row + 1):
        fill = ALT_ROW_FILL if (r_idx - start_row) % 2 == 0 else None
        for c_idx, col in enumerate(columns, start=1):
            key = col.get("key")
            value = get_row_value(row, key, r_idx - start_row)
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if key in ("item_name", "summary", "content", "body", "remark", "action", "decision") else "center", vertical="center", wrap_text=True)
            if re.search(r"price|amount|cost|total|quantity|수량|단가|금액", str(key), re.I):
                num = to_number(value)
                if value not in (None, "") and num:
                    cell.value = int(num) if float(num).is_integer() else num
                    cell.number_format = "#,##0"
            # 행 높이: 내용이 긴 경우
            if value and len(str(value)) > 30:
                ws.row_dimensions[r_idx].height = max(ws.row_dimensions[r_idx].height or 18, 28)
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(columns))}{start_row + len(rows)}"
    ws.freeze_panes = ws.cell(start_row + 1, 1)
    auto_fit_columns(ws)
    apply_print_settings(ws, len(columns), start_row + len(rows), header_row=start_row)

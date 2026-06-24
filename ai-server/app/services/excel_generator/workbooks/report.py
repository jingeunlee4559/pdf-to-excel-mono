from __future__ import annotations

from typing import Any, Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from ..builders import build_meeting_rows, build_official_first_row, build_report_first_row
from ..utils import (
    BORDER,
    DOCUMENT_TYPE_LABELS,
    HEADER_FILL,
    LIGHT_FILL,
    TITLE_FILL,
    as_list,
    get_row_value,
    merge_write,
    set_widths,
    today_text,
    write_cell,
)


def create_report_workbook(payload: Dict[str, Any], doc_type: str) -> Tuple[Workbook, Dict[str, Any]]:
    rows = as_list(payload.get("rows"))
    first = rows[0] if rows else {}
    wb = Workbook()
    ws = wb.active
    ws.title = DOCUMENT_TYPE_LABELS.get(doc_type, "보고서")[:31]

    if doc_type == "MEETING_MINUTES":
        first, action_rows = build_meeting_rows(payload, rows)
        total_cols = 6
        merge_write(ws, 1, 1, 1, total_cols, first.get("meeting_title") or "회 의 록", bold=True, size=18, fill=TITLE_FILL)
        write_cell(ws, 3, 1, "회의일시", bold=True, fill=HEADER_FILL)
        write_cell(ws, 3, 2, first.get("meeting_date") or today_text())
        write_cell(ws, 3, 3, "장소", bold=True, fill=HEADER_FILL)
        write_cell(ws, 3, 4, first.get("meeting_place", ""))
        write_cell(ws, 3, 5, "작성자", bold=True, fill=HEADER_FILL)
        write_cell(ws, 3, 6, payload.get("author_name") or "")
        write_cell(ws, 4, 1, "참석자", bold=True, fill=HEADER_FILL)
        merge_write(ws, 4, 2, 4, total_cols, first.get("attendees", ""))
        sections = [
            ("1. 회의 안건", first.get("agenda") or get_row_value(first, "item_name")),
            ("2. 주요 논의 내용", first.get("discussion") or first.get("content") or ""),
            ("3. 결정 사항", first.get("decision") or ""),
            ("4. 비고", first.get("remark") or ""),
        ]
        r = 6
        for title, body in sections:
            merge_write(ws, r, 1, r, total_cols, title, bold=True, fill=HEADER_FILL)
            merge_write(ws, r + 1, 1, r + 3, total_cols, body or "")
            r += 5
        headers = ["NO", "조치내용", "담당자", "기한", "상태", "비고"]
        for c, h in enumerate(headers, start=1):
            write_cell(ws, r, c, h, bold=True, fill=HEADER_FILL)
        for idx, row in enumerate(action_rows, start=1):
            vals = [idx, row.get("action_item") or "", row.get("owner") or row.get("manager") or "확인 필요", row.get("due_date") or "미정", row.get("status") or "확인 필요", row.get("remark") or ""]
            for c, v in enumerate(vals, start=1):
                write_cell(ws, r + idx, c, v, align="left" if c in (2, 6) else "center")
        set_widths(ws, total_cols, 16)
        ws.column_dimensions["B"].width = 34
        return wb, {"template_kind": doc_type}

    if doc_type == "OFFICIAL_LETTER":
        first = build_official_first_row(payload, rows)
        total_cols = 6
        merge_write(ws, 1, 1, 1, total_cols, first.get("letter_title") or "공 문", bold=True, size=20, fill=TITLE_FILL)
        write_cell(ws, 3, 1, "문서번호", bold=True, fill=HEADER_FILL)
        merge_write(ws, 3, 2, 3, 3, first.get("document_no", ""))
        write_cell(ws, 3, 4, "시행일자", bold=True, fill=HEADER_FILL)
        merge_write(ws, 3, 5, 3, 6, today_text())
        write_cell(ws, 4, 1, "수신", bold=True, fill=HEADER_FILL)
        merge_write(ws, 4, 2, 4, 6, first.get("recipient", ""))
        write_cell(ws, 5, 1, "참조", bold=True, fill=HEADER_FILL)
        merge_write(ws, 5, 2, 5, 6, first.get("reference", ""))
        write_cell(ws, 6, 1, "제목", bold=True, fill=HEADER_FILL)
        merge_write(ws, 6, 2, 6, 6, first.get("document_title") or first.get("title") or "")
        write_cell(ws, 8, 1, "본문", bold=True, fill=HEADER_FILL)
        merge_write(ws, 8, 2, 14, 6, first.get("body") or first.get("content") or first.get("summary") or "")
        write_cell(ws, 16, 1, "붙임", bold=True, fill=HEADER_FILL)
        merge_write(ws, 16, 2, 16, 6, first.get("attachment_note", ""))
        merge_write(ws, 19, 1, 19, 6, first.get("sender") or "공사팀", bold=True, size=14, fill=TITLE_FILL)
        set_widths(ws, total_cols, 16)
        ws.column_dimensions["B"].width = 22
        return wb, {"template_kind": doc_type}

    # BUSINESS REPORT / GENERAL REPORT
    first = build_report_first_row(payload, rows)
    total_cols = 8

    # ── 제목 + 결재란 ───────────────────────────────
    rpt_title = first.get("report_title") or "업무 보고서"
    ws.row_dimensions[1].height = 40
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = ws.cell(1, 1, rpt_title)
    title_cell.font = Font(bold=True, size=18, color="FFFFFF")
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = BORDER
    for c in range(1, 6):
        ws.cell(1, c).border = BORDER

    # 결재란 (우측)
    for c_idx, lbl in enumerate(["담당", "검토", "승인"], start=6):
        ws.row_dimensions[1].height = 40
        write_cell(ws, 1, c_idx, lbl, bold=True, fill=HEADER_FILL)
        ws.row_dimensions[2].height = 36
        write_cell(ws, 2, c_idx, "")  # 서명 공란

    # ── 문서 기본 정보 ──────────────────────────────
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22
    meta_info = [
        ("작성일", today_text()), ("작성자", payload.get("author_name") or ""),
        ("공사명", ""), ("현장명", ""),
    ]
    for i, (lbl, val) in enumerate(meta_info):
        row = 3 + (i // 2)
        col = 1 + (i % 2) * 4
        write_cell(ws, row, col, lbl, bold=True, fill=HEADER_FILL)
        merge_write(ws, row, col + 1, row, col + 3, val, fill=LIGHT_FILL)

    # ── 보고 섹션 ───────────────────────────────────
    section_defs = [
        ("1. 보고 목적", first.get("report_purpose") or "", 3),
        ("2. 주요 검토 내용", first.get("summary") or "", 8),
        ("3. 검토 의견 및 확인사항", first.get("issue_summary") or "", 6),
        ("4. 후속 조치 및 관리계획", first.get("action_plan") or "", 3),
    ]
    r = 6
    for sec_title, body, height in section_defs:
        ws.row_dimensions[r].height = 22
        merge_write(ws, r, 1, r, total_cols, sec_title, bold=True, fill=HEADER_FILL)
        r += 1
        for h in range(height):
            ws.row_dimensions[r + h].height = 20
        merge_write(ws, r, 1, r + height - 1, total_cols, body or "")
        cell = ws.cell(r, 1)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.font = Font(size=10)
        r += height + 1  # 빈 구분 행 1개

    # ── 하단 비고 ────────────────────────────────────
    merge_write(ws, r, 1, r, total_cols, "5. 참고 사항", bold=True, fill=HEADER_FILL)
    ws.row_dimensions[r + 1].height = 20
    merge_write(ws, r + 1, 1, r + 2, total_cols, first.get("footer_note") or "")

    set_widths(ws, total_cols, 16)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    return wb, {"template_kind": doc_type}

from __future__ import annotations

from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..utils import (
    BORDER,
    HEADER_FILL,
    LIGHT_FILL,
    TITLE_FILL,
    merge_write,
    set_widths,
    today_text,
    write_cell,
)

SECTION_FILL = PatternFill("solid", fgColor="2C3E50")   # 섹션 제목 (진한 남색)
SUB_FILL = PatternFill("solid", fgColor="EBF5FB")       # 내용 배경 (연한 파랑)
ACTION_HEADER_FILL = PatternFill("solid", fgColor="1A5276")  # 후속조치 헤더
ALT_FILL = PatternFill("solid", fgColor="F4F6F7")

THIN = Side(style="thin", color="B0BEC5")
MEDIUM = Side(style="medium", color="2C3E50")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
OUTER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

TOTAL_COLS = 8


def _wc(ws, row, col, value, bold=False, size=10, color="000000", fill=None,
        halign="left", valign="top", wrap=True, height=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=size, color=color, name="맑은 고딕")
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = CELL_BORDER
    if fill:
        cell.fill = fill
    if height:
        ws.row_dimensions[row].height = height
    return cell


def _merge(ws, r1, c1, r2, c2, value="", bold=False, size=10,
           color="000000", fill=None, halign="left", valign="top", wrap=True, height=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font = Font(bold=bold, size=size, color=color, name="맑은 고딕")
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = OUTER
    if fill:
        cell.fill = fill
    if height:
        for r in range(r1, r2 + 1):
            ws.row_dimensions[r].height = height
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = CELL_BORDER
    ws.cell(row=r1, column=c1).border = OUTER
    return cell


def create_narrative_report_workbook(
    payload: Dict[str, Any],
    narrative_report: Dict[str, Any],
) -> Tuple[Workbook, Dict[str, Any]]:
    wb = Workbook()
    ws = wb.active
    ws.title = "검토보고서"

    sections = narrative_report.get("sections") or {}
    follow_ups: List[Dict[str, Any]] = narrative_report.get("follow_up_actions") or []
    report_title = str(narrative_report.get("report_title") or "내부 검토보고서")
    author = payload.get("author_name") or ""

    # ── 열 너비 ───────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    r = 1

    # ── 제목 행 ───────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 44
    _merge(ws, r, 1, r, 5, report_title,
           bold=True, size=18, color="FFFFFF", fill=TITLE_FILL,
           halign="center", valign="center")
    # 결재란
    for ci, lbl in enumerate(["담당", "검토", "승인"], start=6):
        _wc(ws, r, ci, lbl, bold=True, size=9, color="FFFFFF",
            fill=SECTION_FILL, halign="center", valign="center", height=44)
    r += 1

    # 결재 서명 공란
    ws.row_dimensions[r].height = 30
    _merge(ws, r, 1, r, 5, "", fill=LIGHT_FILL)
    for ci in range(6, 9):
        _wc(ws, r, ci, "", fill=LIGHT_FILL, height=30)
    r += 1

    # ── 문서 기본 정보 ─────────────────────────────────────────────────
    ws.row_dimensions[r].height = 20
    _wc(ws, r, 1, "작성일", bold=True, fill=SECTION_FILL, color="FFFFFF",
        halign="center", valign="center")
    _merge(ws, r, 2, r, 3, today_text(), fill=SUB_FILL, halign="left")
    _wc(ws, r, 4, "작성자", bold=True, fill=SECTION_FILL, color="FFFFFF",
        halign="center", valign="center")
    _merge(ws, r, 5, r, 6, author, fill=SUB_FILL, halign="left")
    _wc(ws, r, 7, "배포", bold=True, fill=SECTION_FILL, color="FFFFFF",
        halign="center", valign="center")
    _wc(ws, r, 8, "관련 부서", fill=SUB_FILL, halign="left")
    r += 1

    ws.row_dimensions[r].height = 20
    _wc(ws, r, 1, "공사명", bold=True, fill=SECTION_FILL, color="FFFFFF",
        halign="center", valign="center")
    _merge(ws, r, 2, r, 5, "", fill=SUB_FILL, halign="left")
    _wc(ws, r, 6, "현장명", bold=True, fill=SECTION_FILL, color="FFFFFF",
        halign="center", valign="center")
    _merge(ws, r, 7, r, 8, "", fill=SUB_FILL, halign="left")
    r += 1

    # 구분선
    ws.row_dimensions[r].height = 6
    _merge(ws, r, 1, r, TOTAL_COLS, "", fill=TITLE_FILL)
    r += 1

    # ── 보고서 섹션 정의 ────────────────────────────────────────────────
    section_map = [
        ("1. 문서 개요", sections.get("overview") or "원문에서 확인되지 않았습니다. 추가 검토 필요.", 5),
        ("2. 검토 배경 및 목적", sections.get("background") or "원문에서 확인되지 않았습니다. 추가 검토 필요.", 5),
        ("3. 주요 현황", sections.get("current_status") or "원문에서 확인되지 않았습니다. 추가 검토 필요.", 8),
        ("4. 핵심 쟁점", sections.get("key_issues") or "원문에서 확인되지 않았습니다. 추가 검토 필요.", 6),
        ("5. 비용 및 일정 영향", sections.get("cost_schedule_impact") or "원문에서 확인되지 않았습니다. 추가 검토 필요.", 5),
        ("6. 부서별 조치 필요사항", sections.get("department_actions") or "원문에서 확인되지 않았습니다. 추가 검토 필요.", 6),
        ("7. 리스크 및 확인 필요사항", sections.get("risks") or "원문에서 확인되지 않았습니다. 추가 검토 필요.", 5),
        ("8. 종합 검토의견", sections.get("overall_opinion") or "원문에서 확인되지 않았습니다. 추가 검토 필요.", 6),
    ]

    for sec_title, body, lines in section_map:
        # 섹션 헤더
        ws.row_dimensions[r].height = 22
        _merge(ws, r, 1, r, TOTAL_COLS, sec_title,
               bold=True, size=11, color="FFFFFF", fill=SECTION_FILL,
               halign="left", valign="center", wrap=False)
        r += 1

        # 섹션 내용
        content_height = max(lines * 15, 60)
        _merge(ws, r, 1, r + lines - 1, TOTAL_COLS, body,
               size=10, fill=SUB_FILL, halign="left", valign="top", wrap=True)
        for row_idx in range(r, r + lines):
            ws.row_dimensions[row_idx].height = content_height / lines
        r += lines

        # 섹션 간격
        ws.row_dimensions[r].height = 4
        _merge(ws, r, 1, r, TOTAL_COLS, "", fill=PatternFill("solid", fgColor="D5D8DC"))
        r += 1

    # ── 후속 조치사항 테이블 ────────────────────────────────────────────
    ws.row_dimensions[r].height = 24
    _merge(ws, r, 1, r, TOTAL_COLS, "■ 후속 조치사항",
           bold=True, size=12, color="FFFFFF", fill=TITLE_FILL,
           halign="left", valign="center")
    r += 1

    # 테이블 헤더
    ws.row_dimensions[r].height = 20
    headers = ["NO", "담당부서", "조치내용", "목표기한", "우선순위", "상태", "비고", ""]
    col_widths = [4, 12, 28, 12, 8, 8, 8, 0]
    header_fills = ACTION_HEADER_FILL
    for ci, (h, _) in enumerate(zip(headers[:7], col_widths[:7]), start=1):
        _wc(ws, r, ci, h, bold=True, size=10, color="FFFFFF",
            fill=header_fills, halign="center", valign="center")
    _wc(ws, r, 8, "", fill=header_fills)
    r += 1

    if follow_ups:
        for idx, fa in enumerate(follow_ups, start=1):
            ws.row_dimensions[r].height = 20
            row_fill = SUB_FILL if idx % 2 == 0 else ALT_FILL
            _wc(ws, r, 1, idx, halign="center", fill=row_fill)
            _wc(ws, r, 2, fa.get("department") or "확인 필요", fill=row_fill)
            _wc(ws, r, 3, fa.get("action") or "", fill=row_fill)
            _wc(ws, r, 4, fa.get("due_date") or "확인 필요", halign="center", fill=row_fill)
            _wc(ws, r, 5, fa.get("priority") or "보통", halign="center", fill=row_fill)
            _wc(ws, r, 6, fa.get("status") or "미착수", halign="center", fill=row_fill)
            _wc(ws, r, 7, fa.get("remark") or "", fill=row_fill)
            _wc(ws, r, 8, "", fill=row_fill)
            r += 1
    else:
        ws.row_dimensions[r].height = 20
        _merge(ws, r, 1, r, TOTAL_COLS, "후속 조치사항이 확인되지 않았습니다. 추가 검토 필요.",
               fill=SUB_FILL, halign="left")
        r += 1

    # ── 하단 주석 ─────────────────────────────────────────────────────
    ws.row_dimensions[r + 1].height = 16
    _merge(ws, r + 1, 1, r + 1, TOTAL_COLS,
           "※ 본 보고서는 AI 분석 결과를 바탕으로 작성되었습니다. 확인 필요 항목은 담당 부서에서 반드시 검토 후 사용하시기 바랍니다.",
           size=8, color="808080", fill=ALT_FILL, halign="left")

    return wb, {"template_kind": "NARRATIVE_REPORT"}

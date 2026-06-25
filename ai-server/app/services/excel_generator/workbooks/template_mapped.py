from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from ..utils import (
    BASE_DIR,
    as_list,
    get_row_value,
    today_text,
    to_number,
)
from ..vendor_utils import get_vendor_value, infer_vendors, lowest_vendor


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    """src_row의 스타일을 dst_row에 복사 (데이터 행 삽입 시 스타일 유지)"""
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        copy_cell_style(src, dst)
    if ws.row_dimensions.get(src_row):
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _merge_exists(ws, range_text: str) -> bool:
    return any(str(rng) == range_text for rng in ws.merged_cells.ranges)


def _safe_merge(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    if end_row < start_row or end_col < start_col:
        return
    range_text = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    if start_row == end_row and start_col == end_col:
        return
    if _merge_exists(ws, range_text):
        return
    try:
        ws.merge_cells(range_text)
    except Exception:
        pass


def _duplicate_group_merges(ws, merge_specs: List[Tuple[int, int, int, int]], src_start: int, group_width: int, col_offset: int, min_row: int) -> None:
    """마지막 업체 그룹 내부 병합을 새 업체 그룹 위치로 복제한다."""
    src_end = src_start + group_width - 1
    for min_r, min_c, max_r, max_c in merge_specs:
        if min_r < min_row:
            continue
        if min_c >= src_start and max_c <= src_end:
            _safe_merge(ws, min_r, min_c + col_offset, max_r, max_c + col_offset)


def _extend_wide_merges_to_col(ws, original_max_col: int, new_max_col: int) -> None:
    """업체 컬럼 추가 후 제목/최종의견처럼 원래 끝 열까지 넓게 병합된 영역만 새 끝 열까지 확장한다."""
    if not original_max_col or new_max_col <= original_max_col:
        return
    candidates = []
    for rng in list(ws.merged_cells.ranges):
        width = rng.max_col - rng.min_col + 1
        # 전체 폭 성격의 큰 병합만 확장한다.
        # 작성자/견적일자 같은 작은 병합(M4:N4 등)은 확장하지 않는다.
        if rng.max_col == original_max_col and width >= max(6, original_max_col // 2):
            candidates.append((str(rng), rng.min_row, rng.min_col, rng.max_row, new_max_col))
    for old_range, min_r, min_c, max_r, max_c in candidates:
        try:
            ws.unmerge_cells(old_range)
            _safe_merge(ws, min_r, min_c, max_r, max_c)
        except Exception:
            pass


def resolve_template_path(template: Dict[str, Any]) -> Optional[Path]:
    raw = template.get("file_path") or template.get("filePath") or ""
    if not raw:
        return None
    candidates = [
        Path(raw),
        Path.cwd() / raw,
        BASE_DIR / raw,
        BASE_DIR.parent / raw,
        Path(raw.replace("\\", "/")),
        Path(raw.replace("/", "\\")),
    ]
    for candidate in candidates:
        try:
            if candidate.exists():
                return candidate
        except Exception:
            pass
    return None


def cell_col(cell_address: str) -> Optional[int]:
    m = re.match(r"^([A-Z]+)", str(cell_address or ""), re.I)
    return column_index_from_string(m.group(1).upper()) if m else None


def cell_row_num(cell_address: str) -> Optional[int]:
    m = re.search(r"(\d+)$", str(cell_address or ""))
    return int(m.group(1)) if m else None


def _resolve_single_cell_value(field_key: str, payload: Dict[str, Any]) -> Any:
    """SINGLE_CELL 필드 값을 여러 소스에서 탐색하여 반환"""
    rows = as_list(payload.get("rows"))
    analysis = payload.get("analysis") or {}
    job = payload.get("job") or {}
    table_json = {}
    if isinstance(job.get("tables"), list) and job["tables"]:
        table_json = job["tables"][0].get("tableJson") or job["tables"][0].get("table_json") or {}

    # 1. 날짜 필드
    if field_key in ("document_date", "date", "created_date", "견적일자", "작성일"):
        return today_text()

    # 2. 작성자 필드
    if field_key in ("requester_name", "writer_name", "created_by", "author", "작성자"):
        return payload.get("author_name") or payload.get("authorName") or ""

    # 3. 문서명/제목 필드
    if field_key in ("document_title", "title", "문서명", "제목"):
        job_user_req = job.get("userRequest") or job.get("user_request") or ""
        doc_type = analysis.get("documentType") or analysis.get("document_type") or ""
        table_name = job.get("tables", [{}])[0].get("tableName") if isinstance(job.get("tables"), list) and job.get("tables") else ""
        return table_name or doc_type or "비교견적서" or ""

    # 4. 보고서 섹션 필드 (narrativeReport에서 추출)
    narrative = analysis.get("narrativeReport") or analysis.get("narrative_report") or {}
    if isinstance(narrative, dict):
        sections = narrative.get("sections") or {}
        section_map = {
            "overview": sections.get("overview", ""),
            "background": sections.get("background", ""),
            "current_status": sections.get("current_status", ""),
            "key_issues": sections.get("key_issues", ""),
            "overall_opinion": sections.get("overall_opinion", ""),
            "department_actions": sections.get("department_actions", ""),
            "risks": sections.get("risks", ""),
            "cost_schedule_impact": sections.get("cost_schedule_impact", ""),
            "final_opinion": sections.get("overall_opinion", ""),
            "special_note": sections.get("key_issues", ""),
        }
        if field_key in section_map and section_map[field_key]:
            return section_map[field_key]

    # 5. analysis.keyValues에서 탐색
    key_values = analysis.get("keyValues") or []
    for kv in key_values:
        label = str(kv.get("label") or "").lower()
        if field_key.lower() in label or label in field_key.lower():
            return kv.get("value") or ""

    # 6. table_json 메타데이터에서 탐색
    meta = table_json.get("meta") or {}
    if field_key in meta:
        return meta[field_key]

    # 7. rows[0]에서 탐색 (마지막 수단)
    if rows:
        val = get_row_value(rows[0], field_key)
        if val not in (None, "", 0):
            return val

    # 8. 제한적 fallback — 기타사항/최종의견에는 분석 요약을 자동 기입하지 않음
    # (사용자가 직접 입력하지 않은 AI 텍스트가 자사 양식에 들어가는 문제 방지)
    fallback_map = {
        "project_name": job.get("title") or job.get("userRequest") or "",
        "site_name": job.get("title") or "",
    }
    if field_key in fallback_map:
        return fallback_map[field_key]

    return ""



def _vendor_unit_prices(row: Dict[str, Any], vendors: List[Dict[str, Any]]) -> List[Tuple[Dict[str, Any], float]]:
    """행 기준 업체별 단가 숫자 목록을 반환한다."""
    prices: List[Tuple[Dict[str, Any], float]] = []
    for vendor in vendors or []:
        price = to_number(get_vendor_value(row, vendor, "unit_price"))
        if price and price > 0:
            prices.append((vendor, float(price)))
    return prices


def _calculated_average_price(row: Dict[str, Any], vendors: List[Dict[str, Any]]) -> Any:
    """업체별 제품가격 조사현황표의 평균가격 fallback."""
    explicit = get_row_value(row, "average_price") or get_row_value(row, "avg_price") or get_row_value(row, "average_unit_price")
    if explicit not in (None, ""):
        return explicit
    prices = [price for _, price in _vendor_unit_prices(row, vendors)]
    if not prices:
        return ""
    avg = sum(prices) / len(prices)
    return int(round(avg)) if float(avg).is_integer() or avg >= 1 else avg


def _calculated_selected_vendor(row: Dict[str, Any], vendors: List[Dict[str, Any]]) -> Any:
    """업체선정 fallback: 명시값이 없으면 최저 단가 업체명을 반환한다."""
    explicit = (
        get_row_value(row, "selected_vendor") or
        get_row_value(row, "selected_company") or
        get_row_value(row, "chosen_vendor") or
        get_row_value(row, "lowest_vendor") or
        get_row_value(row, "lowest_target")
    )
    if explicit not in (None, ""):
        return explicit
    name, _price = lowest_vendor(row, vendors)
    return name or ""


def _clear_cell_if_not_in_range(ws, row: int, col: int, keep_min_col: int, keep_max_col: int) -> None:
    if keep_min_col <= col <= keep_max_col:
        return
    try:
        ws.cell(row, col).value = None
    except Exception:
        pass


def _unmerge_intersecting_row(ws, row: int, min_col: int, max_col: int) -> None:
    """특정 행/열 범위와 겹치는 병합을 해제한다."""
    targets = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row and not (rng.max_col < min_col or rng.min_col > max_col):
            targets.append(str(rng))
    for range_text in targets:
        try:
            ws.unmerge_cells(range_text)
        except Exception:
            pass


def _find_cell_contains(ws, keywords: List[str], max_scan_row: int = 8) -> Optional[Tuple[int, int]]:
    for row in range(1, min(ws.max_row or 1, max_scan_row) + 1):
        for col in range(1, (ws.max_column or 1) + 1):
            value = str(ws.cell(row, col).value or "").replace(" ", "")
            if value and any(keyword.replace(" ", "") in value for keyword in keywords):
                return row, col
    return None


def _move_author_block_to_right_edge(ws, payload: Dict[str, Any], company_group_info: Optional[Dict[str, Any]]) -> None:
    """비교견적서 양식에서 작성자 블록을 실제 업체 컬럼 오른쪽 끝으로 이동한다.

    업체가 4개 이상으로 늘면 기존 K:N 작성자 영역이 업체 컬럼과 겹친다.
    따라서 최종 테이블 우측 4칸을 작성자 라벨/값 영역으로 재배치한다.
    """
    if not company_group_info:
        return
    try:
        gw = int(company_group_info.get("group_width") or 0)
        real_cnt = int(company_group_info.get("actual_vendor_count") or 0)
        tmpl_cnt = int(company_group_info.get("template_vendor_count") or 0)
        letters = as_list(company_group_info.get("letters"))
    except Exception:
        return
    # 제품가격 조사현황표(group_width=1)는 대상이 아니다.
    # 작성자 블록 이동은 업체 수가 원본 업체 슬롯보다 많아져 기존 작성자 영역과 충돌할 때만 수행한다.
    if gw < 4 or real_cnt <= tmpl_cnt:
        return

    found = _find_cell_contains(ws, ["작성자", "작성"])
    if not found:
        return
    author_row, old_label_col = found

    # 실제 업체 컬럼 오른쪽 끝. 기본 컬럼 A:B + 업체그룹(4칸×업체수)을 기준으로 계산하고,
    # 동적 삽입 후 ws.max_column이 더 넓으면 그 값을 우선한다.
    try:
        first_vendor_col = column_index_from_string(letters[0]) if letters else 3
    except Exception:
        first_vendor_col = 3
    table_right_col = max(ws.max_column or 1, first_vendor_col - 1 + real_cnt * gw)
    if table_right_col < 4:
        return

    label_start = table_right_col - 3
    label_end = table_right_col - 2
    value_start = table_right_col - 1
    value_end = table_right_col
    if label_start < 1:
        return

    old_label_cell = ws.cell(author_row, old_label_col)
    old_value_col = old_label_col + 2
    old_value_cell = ws.cell(author_row, old_value_col) if old_value_col <= (ws.max_column or old_value_col) else old_label_cell
    old_label = old_label_cell.value or "작  성  자"
    author_name = payload.get("author_name") or payload.get("authorName") or old_value_cell.value or ""

    # 도착 범위 병합 충돌 제거 후 작성자 라벨/값 병합 생성
    _unmerge_intersecting_row(ws, author_row, label_start, value_end)
    # 기존 작성자 병합도 도착지와 다르면 해제/정리
    _unmerge_intersecting_row(ws, author_row, old_label_col, min(ws.max_column or old_label_col, old_label_col + 3))

    for col in range(old_label_col, min(ws.max_column or old_label_col, old_label_col + 4) + 1):
        _clear_cell_if_not_in_range(ws, author_row, col, label_start, value_end)

    try:
        _safe_merge(ws, author_row, label_start, author_row, label_end)
        _safe_merge(ws, author_row, value_start, author_row, value_end)
    except Exception:
        pass

    for col in range(label_start, label_end + 1):
        copy_cell_style(old_label_cell, ws.cell(author_row, col))
    for col in range(value_start, value_end + 1):
        copy_cell_style(old_value_cell, ws.cell(author_row, col))

    ws.cell(author_row, label_start).value = old_label
    ws.cell(author_row, value_start).value = author_name

    # 새 작성자 영역은 숨김 컬럼이면 보이지 않으므로 반드시 표시한다.
    for col in range(label_start, value_end + 1):
        try:
            letter = get_column_letter(col)
            ws.column_dimensions[letter].hidden = False
            if not ws.column_dimensions[letter].width or ws.column_dimensions[letter].width < 4:
                ws.column_dimensions[letter].width = 10
        except Exception:
            pass


def _merge_quote_date_until_author(ws, payload: Dict[str, Any]) -> None:
    """비교견적서 상단의 견적일자 값 영역을 작성자 블록 직전까지 병합한다.

    업체 컬럼이 동적으로 늘어나면 작성자 블록은 우측 끝으로 이동하지만,
    기존 견적일자 값 영역(I:J)이 그대로 남아 중간에 빈 셀이 생긴다.
    예: G:H=견적일자, I:N=날짜, O:P=작성자, Q:R=작성자명
    """
    date_found = _find_cell_contains(ws, ["견적일자", "견적 일자", "작성일자", "작성 일자"], max_scan_row=8)
    author_found = _find_cell_contains(ws, ["작성자", "작  성  자", "작 성 자"], max_scan_row=8)
    if not date_found or not author_found:
        return

    date_row, date_col = date_found
    author_row, author_col = author_found
    if date_row != author_row:
        # 같은 행에서 작성자 셀을 다시 찾는다. 다른 행의 작성자 문구를 잘못 잡는 것을 방지한다.
        for col in range(1, (ws.max_column or 1) + 1):
            text = str(ws.cell(date_row, col).value or "").replace(" ", "")
            if text in ("작성자", "작성") or "작성자" in text:
                author_row, author_col = date_row, col
                break
        if date_row != author_row:
            return

    date_label_start = date_col
    date_label_end = date_col
    author_start = author_col

    # 병합 범위 기준으로 라벨 끝/작성자 시작 위치를 정확히 잡는다.
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= date_row <= rng.max_row and rng.min_col <= date_col <= rng.max_col:
            date_label_start = rng.min_col
            date_label_end = rng.max_col
        if rng.min_row <= author_row <= rng.max_row and rng.min_col <= author_col <= rng.max_col:
            author_start = rng.min_col

    value_start = date_label_end + 1
    value_end = author_start - 1
    if value_end < value_start:
        return

    # 기존 날짜값 보존. 없으면 payload/오늘 날짜 fallback.
    date_value = None
    for col in range(value_start, value_end + 1):
        val = ws.cell(date_row, col).value
        if val not in (None, ""):
            date_value = val
            break
    if date_value in (None, ""):
        date_value = (
            payload.get("document_date") or
            payload.get("date") or
            payload.get("created_date") or
            today_text()
        )

    style_source = ws.cell(date_row, value_start)
    if not style_source.has_style and value_start - 1 >= 1:
        style_source = ws.cell(date_row, value_start - 1)

    _unmerge_intersecting_row(ws, date_row, value_start, value_end)
    for col in range(value_start, value_end + 1):
        cell = ws.cell(date_row, col)
        copy_cell_style(style_source, cell)
        cell.value = None

    _safe_merge(ws, date_row, value_start, date_row, value_end)
    target = ws.cell(date_row, value_start)
    target.value = date_value
    copy_cell_style(style_source, target)

    for col in range(value_start, value_end + 1):
        try:
            letter = get_column_letter(col)
            ws.column_dimensions[letter].hidden = False
            if not ws.column_dimensions[letter].width or ws.column_dimensions[letter].width < 4:
                ws.column_dimensions[letter].width = 10
        except Exception:
            pass

def _insert_rows_for_repeat(ws, start_row: int, template_rows: int, needed_rows: int, max_col: int) -> int:
    """
    템플릿에 할당된 행(template_rows)보다 실제 데이터(needed_rows)가 많을 때
    부족한 행 수만큼 삽입하고, 삽입한 행에 스타일 복사.
    반환값: 실제 삽입된 행 수
    """
    if needed_rows <= template_rows:
        return 0
    insert_count = needed_rows - template_rows
    insert_at = start_row + template_rows  # 데이터 영역 바로 다음 행에 삽입
    ws.insert_rows(insert_at, amount=insert_count)
    # 마지막 템플릿 데이터 행의 스타일을 새 행에 복사
    style_src_row = start_row + template_rows - 1
    for new_row in range(insert_at, insert_at + insert_count):
        copy_row_style(ws, style_src_row, new_row, max_col)
    return insert_count


def create_mapped_template_workbook(payload: Dict[str, Any]) -> Optional[Tuple[Workbook, Dict[str, Any]]]:
    template = payload.get("template") or {}
    mapping_json = payload.get("mapping_json") or {}
    mappings: List[Dict[str, Any]] = as_list(payload.get("mappings") or mapping_json.get("mappings"))
    template_path = resolve_template_path(template)

    if not template_path:
        return None
    if not mappings:
        return None

    try:
        wb = load_workbook(template_path)
    except Exception as e:
        return None

    # 시트 선택
    sheet_name = mappings[0].get("sheetName") if mappings else None
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    rows = as_list(payload.get("rows"))
    columns = as_list(payload.get("columns"))
    job = payload.get("job") or {}
    analysis = payload.get("analysis") or {}
    table_json = {}
    if isinstance(job.get("tables"), list) and job["tables"]:
        table_json = job["tables"][0].get("tableJson") or job["tables"][0].get("table_json") or {}
    # 다운로드/미리보기 엑셀 생성 시 컬럼 라벨만으로 업체명을 추론하면
    # "5개 업체 단가 비교" 같은 표 제목이 업체명으로 들어갈 수 있다.
    # 원 채팅 요청을 meta에 주입해 infer_vendors가 실제 요청 업체 순서를 우선 사용하도록 한다.
    if isinstance(table_json, dict):
        meta = table_json.get("meta") if isinstance(table_json.get("meta"), dict) else {}
        table_json = {
            **table_json,
            "meta": {
                **meta,
                "userRequest": meta.get("userRequest") or meta.get("user_request") or job.get("userRequest") or job.get("user_request") or job.get("title") or "",
            },
        }

    vendors = infer_vendors(columns, rows, table_json)

    # table_json.meta.vendors가 없으면 analysis.fileProfiles에서 업체명 보완
    if not vendors:
        file_profiles = analysis.get("fileProfiles") or analysis.get("file_profiles") or []
        for idx, profile in enumerate(file_profiles):
            if not isinstance(profile, dict):
                continue
            name = str(
                profile.get("companyName") or profile.get("company_name") or
                profile.get("vendorName") or profile.get("vendor_name") or ""
            ).strip()
            if name:
                vendors.append({"name": name, "index": idx})

    max_col = ws.max_column or 14
    original_max_col = max_col

    # ── 1단계: REPEAT_ROW 매핑에서 데이터 영역 파악 및 행 삽입 ──────────────
    repeat_mappings = [m for m in mappings if str(m.get("mappingType") or m.get("mapping_type") or "").upper() == "REPEAT_ROW"]
    if repeat_mappings and rows:
        rep = repeat_mappings[0]
        start_row = int(rep.get("startRow") or rep.get("start_row") or 7)
        template_max_rows = int(rep.get("maxRows") or rep.get("max_rows") or 16)
        needed = len(rows)
        inserted = _insert_rows_for_repeat(ws, start_row, template_max_rows, needed, max_col)
    else:
        inserted = 0

    # ── 2단계: 각 매핑 처리 ──────────────────────────────────────────────────
    company_group_info: Optional[Dict[str, Any]] = None  # 업체 컬럼 후처리용
    for m in mappings:
        field_key = m.get("fieldKey") or m.get("field_key") or ""
        mapping_type = str(m.get("mappingType") or m.get("mapping_type") or "").upper()

        # ── SINGLE_CELL ───────────────────────────────────────────────────────
        if mapping_type == "SINGLE_CELL":
            addr = m.get("cellAddress") or str(m.get("mergedRange") or "").split(":")[0]
            if not addr:
                continue
            # 행 삽입 후 주소 재계산 (footer 행들은 밀려남)
            orig_row = cell_row_num(addr)
            if orig_row and inserted > 0:
                rep_start = int(repeat_mappings[0].get("startRow") or 7) if repeat_mappings else 99
                rep_end = rep_start + int(repeat_mappings[0].get("maxRows") or 16) - 1 if repeat_mappings else 99
                if orig_row > rep_end:
                    col_letter = re.match(r"^([A-Z]+)", addr, re.I)
                    if col_letter:
                        addr = f"{col_letter.group(1)}{orig_row + inserted}"

            val = _resolve_single_cell_value(field_key, payload)
            try:
                ws[addr] = val
            except Exception:
                pass

        # ── REPEAT_ROW ────────────────────────────────────────────────────────
        elif mapping_type in ("REPEAT_ROW", "REPEAT_COLUMN"):
            col_letter = m.get("columnLetter") or re.sub(r"\d", "", str(m.get("cellAddress") or "")).strip()
            if not col_letter:
                continue
            try:
                col = column_index_from_string(str(col_letter).upper())
            except Exception:
                continue
            start = int(m.get("startRow") or m.get("start_row") or cell_row_num(m.get("cellAddress")) or 7)
            write_count = len(rows)
            for idx in range(write_count):
                if idx < len(rows) and field_key in ("average_price", "avg_price", "average_unit_price", "평균가격", "평균단가"):
                    val = _calculated_average_price(rows[idx], vendors)
                elif idx < len(rows) and field_key in ("selected_vendor", "selected_company", "chosen_vendor", "lowest_vendor", "lowest_target", "업체선정", "최저업체"):
                    val = _calculated_selected_vendor(rows[idx], vendors)
                else:
                    val = get_row_value(rows[idx], field_key, idx) if idx < len(rows) else None
                if val not in (None, ""):
                    try:
                        cell = ws.cell(start + idx, col)
                        cell.value = val
                        # 숫자 서식 적용
                        if field_key in ("quantity", "unit_price", "amount", "total_amount", "average_price", "avg_price", "average_unit_price", "평균가격", "평균단가") and to_number(val):
                            cell.value = int(to_number(val)) if float(to_number(val)).is_integer() else to_number(val)
                            cell.number_format = "#,##0"
                    except Exception:
                        pass

        # ── COMPANY_GROUP_COLUMN ──────────────────────────────────────────────
        elif mapping_type == "COMPANY_GROUP_COLUMN":
            start = int(m.get("startRow") or m.get("start_row") or 7)
            field = field_key
            letters: List[str] = as_list(m.get("columnLetters") or m.get("column_letters"))
            group_width = int(m.get("groupWidth") or m.get("group_width") or 4)
            group_ranges: List[str] = as_list(m.get("groupRanges") or m.get("group_ranges"))

            if field in ("target_name", "vendor_name", "company_name"):
                template_vendor_count = max(len(group_ranges), len(letters)) if (group_ranges or letters) else 1
                actual_vendor_count = len(vendors)

                # 헤더 행 번호 파악 (group_ranges 또는 start-1)
                header_row: Optional[int] = None
                if group_ranges and group_ranges[0]:
                    header_row = cell_row_num(str(group_ranges[0]).split(":")[0])
                if header_row is None:
                    header_row = max(1, start - 1)

                # ① 실제 업체 수만큼 헤더 기록 (템플릿 슬롯 범위 내)
                for idx, vendor in enumerate(vendors):
                    if idx >= template_vendor_count:
                        break
                    vendor_name = vendor.get("name", "")
                    # group_ranges 방식 (등록된 헤더 셀 주소 우선)
                    if idx < len(group_ranges) and group_ranges[idx]:
                        header_addr = str(group_ranges[idx]).split(":")[0]
                        try:
                            ws[header_addr] = vendor_name
                        except Exception:
                            pass
                    # letters+header_row 방식 (보완 기록 — group_ranges 주소가 틀렸을 경우도 커버)
                    if letters and idx < len(letters):
                        try:
                            ws[f"{letters[idx]}{header_row}"] = vendor_name
                        except Exception:
                            pass

                # ② 초과 업체: 오른쪽에 컬럼 그룹 삽입 (데이터 기록 전에 실행)
                if actual_vendor_count > template_vendor_count and letters:
                    ref_start = column_index_from_string(letters[-1])
                    extra_count = actual_vendor_count - template_vendor_count

                    # 업체 데이터가 실제로 시작되는 행과, 스타일을 유지해야 할 마지막 행을 분리한다.
                    # target_name의 startRow는 헤더 행이라서 이를 데이터 시작으로 쓰면
                    # 새 업체 그룹의 "규격/수량/단가/금액" 서브헤더가 비어 버린다.
                    group_data_start_row: Optional[int] = None
                    group_style_end_row = ws.max_row or start
                    for gm in mappings:
                        if str(gm.get("mappingType") or gm.get("mapping_type") or "").upper() != "COMPANY_GROUP_COLUMN":
                            continue
                        try:
                            gm_field = str(gm.get("fieldKey") or gm.get("field_key") or "")
                            gm_start = int(gm.get("startRow") or gm.get("start_row") or start)
                            gm_end = int(gm.get("endRow") or gm.get("end_row") or (gm_start + int(gm.get("maxRows") or gm.get("max_rows") or 1) - 1))
                            group_style_end_row = max(group_style_end_row, gm_end)
                            if gm_field not in ("target_name", "vendor_name", "company_name") and gm_start > header_row:
                                group_data_start_row = gm_start if group_data_start_row is None else min(group_data_start_row, gm_start)
                        except Exception:
                            pass
                    if group_data_start_row is None:
                        group_data_start_row = header_row + 2

                    merge_specs = [(rng.min_row, rng.min_col, rng.max_row, rng.max_col) for rng in list(ws.merged_cells.ranges)]

                    for ex in range(extra_count):
                        insert_at = ref_start + group_width * (ex + 1)
                        col_offset = group_width * (ex + 1)
                        try:
                            ws.insert_cols(insert_at, group_width)
                        except Exception:
                            continue

                        # 마지막 템플릿 업체 컬럼의 너비·스타일·서브헤더·하단 빈 양식까지 복사
                        for j in range(group_width):
                            try:
                                src_letter = get_column_letter(ref_start + j)
                                dst_letter = get_column_letter(insert_at + j)
                                ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
                                ws.column_dimensions[dst_letter].hidden = False
                            except Exception:
                                pass
                        for row_num in range(header_row, group_style_end_row + 1):
                            for j in range(group_width):
                                try:
                                    src = ws.cell(row_num, ref_start + j)
                                    dst = ws.cell(row_num, insert_at + j)
                                    copy_cell_style(src, dst)
                                    # 업체 헤더/서브헤더는 복사하고, 실제 데이터 행은 값이 섞이지 않도록 비운다.
                                    dst.value = src.value if row_num < group_data_start_row else None
                                except Exception:
                                    pass

                        # 마지막 업체 그룹 내부의 병합 영역도 새 업체 그룹에 복제한다.
                        _duplicate_group_merges(ws, merge_specs, ref_start, group_width, col_offset, header_row)

                        # 업체명 헤더 셀: 병합 + 스타일 + 업체명 기록
                        vendor_idx = template_vendor_count + ex
                        if vendor_idx < actual_vendor_count:
                            vendor_name = vendors[vendor_idx].get("name", "")
                            try:
                                hl = get_column_letter(insert_at)
                                hr_l = get_column_letter(insert_at + group_width - 1)
                                if group_width > 1:
                                    _safe_merge(ws, header_row, insert_at, header_row, insert_at + group_width - 1)
                                hcell = ws.cell(header_row, insert_at)
                                copy_cell_style(ws.cell(header_row, ref_start), hcell)
                                hcell.value = vendor_name
                            except Exception:
                                pass

                # 후처리(삭제·너비·A4)를 위해 정보 저장
                company_group_info = {
                    "letters": letters[:],
                    "group_width": group_width,
                    "template_vendor_count": template_vendor_count,
                    "actual_vendor_count": actual_vendor_count,
                }
                continue

            # 데이터 행 기록 — 동적 업체 컬럼 처리
            write_count = len(rows)
            for vidx, vendor in enumerate(vendors):
                # 컬럼 위치 결정: 템플릿 정의 범위 내 / 초과 업체 처리
                if letters and vidx < len(letters):
                    try:
                        col = column_index_from_string(letters[vidx])
                    except Exception:
                        continue
                elif letters:
                    # 초과 업체: 마지막 정의 컬럼 이후 group_width 단위로 배치
                    try:
                        last_col = column_index_from_string(letters[-1])
                        col = last_col + (vidx - len(letters) + 1) * group_width
                    except Exception:
                        continue
                else:
                    col = 6 + vidx * group_width

                for ridx in range(write_count):
                    val = get_vendor_value(rows[ridx], vendor, field) if ridx < len(rows) else None
                    if val not in (None, ""):
                        try:
                            cell = ws.cell(start + ridx, col)
                            cell.value = val
                            if field in ("unit_price", "amount", "total_amount") and to_number(val):
                                cell.value = int(to_number(val)) if float(to_number(val)).is_integer() else to_number(val)
                                cell.number_format = "#,##0"
                        except Exception:
                            pass

    # ── 3단계: 미사용 업체 컬럼 삭제 + 너비 재분배 + A4 설정 ─────────────────
    # target_name 매핑이 없는 템플릿(예: 제품가격 조사현황표)도 처리하도록
    # company_group_info가 없으면 다른 COMPANY_GROUP_COLUMN 매핑에서 복구 시도
    if company_group_info is None:
        for _m in mappings:
            if str(_m.get("mappingType") or _m.get("mapping_type") or "").upper() == "COMPANY_GROUP_COLUMN":
                _letters = as_list(_m.get("columnLetters") or _m.get("column_letters"))
                if _letters:
                    _gw = int(_m.get("groupWidth") or _m.get("group_width") or 1)
                    _gr = as_list(_m.get("groupRanges") or _m.get("group_ranges"))
                    _tmpl = max(len(_gr), len(_letters)) if (_gr or _letters) else 1
                    company_group_info = {
                        "letters": _letters,
                        "group_width": _gw,
                        "template_vendor_count": _tmpl,
                        "actual_vendor_count": len(vendors),
                    }
                    break

    if company_group_info:
        letters_all: List[str] = company_group_info["letters"]
        gw: int = company_group_info["group_width"]
        tmpl_cnt: int = company_group_info["template_vendor_count"]
        real_cnt: int = company_group_info["actual_vendor_count"]

        if real_cnt > 0 and real_cnt < tmpl_cnt and letters_all:
            # 미사용 업체 그룹은 삭제하지 않는다.
            # delete_cols()는 오른쪽 비고/합계/수식/병합 좌표를 밀어 자사양식을 깨뜨린다.
            # 대신 숨김 처리 + 값 제거로 브라우저/다운로드 모두에서 빈 업체 슬롯이 보이지 않게 한다.
            hidden_col_indexes: List[int] = []
            for i in range(real_cnt, tmpl_cnt):
                if i < len(letters_all):
                    try:
                        ci = column_index_from_string(letters_all[i])
                        for j in range(gw):
                            col_idx = ci + j
                            hidden_col_indexes.append(col_idx)
                            cl = get_column_letter(col_idx)
                            ws.column_dimensions[cl].hidden = True
                            ws.column_dimensions[cl].width = 0.1
                            for rr in range(1, (ws.max_row or 1) + 1):
                                try:
                                    ws.cell(rr, col_idx).value = None
                                except Exception:
                                    pass
                    except Exception:
                        pass

            # 평균/합계 수식이 숨김 업체 끝 열까지 잡혀 있으면 실제 업체 끝 열로 축소한다.
            try:
                first_vendor_col = column_index_from_string(letters_all[0])
                visible_last_start = column_index_from_string(letters_all[real_cnt - 1])
                template_last_start = column_index_from_string(letters_all[tmpl_cnt - 1])
                visible_last_col = visible_last_start + gw - 1
                template_last_col = template_last_start + gw - 1
                old_end_letter = get_column_letter(template_last_col)
                new_end_letter = get_column_letter(visible_last_col)
                first_letter = get_column_letter(first_vendor_col)
                if template_last_col != visible_last_col:
                    formula_re = re.compile(rf"({first_letter}\$?\d+):({old_end_letter})(\$?\d+)")
                    for row_cells in ws.iter_rows():
                        for cell in row_cells:
                            if isinstance(cell.value, str) and cell.value.startswith("="):
                                cell.value = formula_re.sub(lambda mm: f"{mm.group(1)}:{new_end_letter}{mm.group(3)}", cell.value)
            except Exception:
                pass

    # ── 4단계: 비교견적서 작성자 블록을 실제 표 오른쪽 끝으로 재배치 ─────────
    try:
        _move_author_block_to_right_edge(ws, payload, company_group_info)
    except Exception:
        pass

    # ── 4-1단계: 견적일자 값 영역을 작성자 직전까지 병합 ───────────────────
    try:
        _merge_quote_date_until_author(ws, payload)
    except Exception:
        pass

    # ── 5단계: 제목/최종의견처럼 전체 폭 성격의 병합만 새 max_column까지 확장 ───
    try:
        _extend_wide_merges_to_col(ws, original_max_col, ws.max_column or original_max_col)
    except Exception:
        pass

    # A4 페이지 자동 맞춤
    try:
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetPr.fitToPage = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.35
        ws.page_margins.bottom = 0.35
        ws.print_options.horizontalCentered = True
        ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    except Exception:
        pass

    return wb, {"template_kind": "MAPPED_COMPANY_TEMPLATE", "vendor_count": len(vendors)}

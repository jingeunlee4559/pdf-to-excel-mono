from __future__ import annotations

from typing import Any, Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..utils import (
    ALT_ROW_FILL,
    BORDER,
    GREEN_FILL,
    HEADER_FILL,
    HEADER_FILL2,
    LIGHT_FILL,
    TITLE_FILL,
    as_list,
    get_row_value,
    merge_write,
    set_widths,
    to_number,
    today_text,
    write_cell,
    write_table,
)
from ..vendor_utils import (
    comparable_company,
    get_vendor_value,
    infer_vendors,
    lowest_vendor,
)


def create_estimate_comparison_workbook(payload: Dict[str, Any]) -> Tuple[Workbook, Dict[str, Any]]:
    rows = as_list(payload.get("rows"))
    columns = as_list(payload.get("columns"))
    job = payload.get("job") or {}
    table_json = {}
    if isinstance(job.get("tables"), list) and job["tables"]:
        table_json = job["tables"][0].get("tableJson") or job["tables"][0].get("table_json") or {}
    vendors = infer_vendors(columns, rows, table_json)
    if not vendors:
        vendors = [{"name": "업체1", "index": 0}, {"name": "업체2", "index": 1}]

    # 표준시장단가 포함 여부 판단
    has_std = any(to_number(row.get("standard_unit_price")) > 0 for row in rows)

    base_cols = [
        {"key": "row_no", "label": "NO"},
        {"key": "construction_code", "label": "공종코드"},
        {"key": "item_name", "label": "품목명"},
        {"key": "spec", "label": "규격"},
        {"key": "quantity", "label": "수량"},
        {"key": "unit", "label": "단위"},
    ]
    # 공종코드 없으면 제거
    if not any(row.get("construction_code") for row in rows):
        base_cols = [c for c in base_cols if c["key"] != "construction_code"]

    summary_cols = 3  # 최저업체, 최저단가, 비고
    std_extra = 1 if has_std else 0
    total_cols = len(base_cols) + std_extra + len(vendors) * 2 + summary_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "업체별단가비교"

    # ── 타이틀 ──────────────────────────────────────
    doc_title = str((job.get("analysis") or {}).get("documentType") or "업체별 단가 비교표")
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(1, 1, doc_title)
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = BORDER

    # ── 메타 정보 행 ─────────────────────────────────
    # 업체 수가 늘어 total_cols가 커질 때도 작성자 영역이 표의 오른쪽 끝에 붙도록 배치한다.
    # 기존 방식(idx * 3)은 작성자/시스템관리자가 왼쪽에 고정되어 우측이 빈칸으로 남았다.
    ws.row_dimensions[2].height = 20
    author = payload.get("author_name") or ""
    if total_cols >= 8:
        write_cell(ws, 2, 1, "작성일", bold=True, fill=HEADER_FILL)
        date_end = min(4, max(2, total_cols - 4))
        merge_write(ws, 2, 2, 2, date_end, today_text(), fill=LIGHT_FILL)
        # 가운데 빈 구간도 양식 배경/테두리 유지
        for c in range(date_end + 1, max(date_end + 1, total_cols - 2)):
            write_cell(ws, 2, c, "", fill=LIGHT_FILL)
        write_cell(ws, 2, total_cols - 2, "작성자", bold=True, fill=HEADER_FILL)
        merge_write(ws, 2, total_cols - 1, 2, total_cols, author, fill=LIGHT_FILL)
    else:
        meta_pairs = [("작성일", today_text()), ("작성자", author)]
        for idx, (lbl, val) in enumerate(meta_pairs):
            col = idx * 3 + 1
            if col > total_cols:
                break
            write_cell(ws, 2, col, lbl, bold=True, fill=HEADER_FILL)
            merge_write(ws, 2, min(col + 1, total_cols), 2, min(col + 2, total_cols), val, fill=LIGHT_FILL)

    # ── 헤더 (2단) ──────────────────────────────────
    header_row = 3
    ws.row_dimensions[header_row].height = 22
    ws.row_dimensions[header_row + 1].height = 20

    for idx, col in enumerate(base_cols, start=1):
        merge_write(ws, header_row, idx, header_row + 1, idx, col["label"], bold=True, fill=HEADER_FILL)

    col_cursor = len(base_cols) + 1

    if has_std:
        merge_write(ws, header_row, col_cursor, header_row + 1, col_cursor, "표준단가", bold=True, fill=HEADER_FILL2)
        col_cursor += 1

    # 업체별 2칸 (단가/금액)
    vendor_fill_cycle = [HEADER_FILL, HEADER_FILL2]
    for v_idx, vendor in enumerate(vendors):
        vname = vendor.get("name") or f"업체{v_idx + 1}"
        vfill = vendor_fill_cycle[v_idx % 2]
        merge_write(ws, header_row, col_cursor, header_row, col_cursor + 1, vname, bold=True, fill=vfill)
        write_cell(ws, header_row + 1, col_cursor, "단가", bold=True, fill=LIGHT_FILL)
        write_cell(ws, header_row + 1, col_cursor + 1, "금액", bold=True, fill=LIGHT_FILL)
        col_cursor += 2

    for lbl in ["최저 업체", "최저 단가", "비고"]:
        merge_write(ws, header_row, col_cursor, header_row + 1, col_cursor, lbl, bold=True, fill=HEADER_FILL)
        col_cursor += 1

    # ── 데이터 행 ────────────────────────────────────
    data_start = header_row + 2
    for row_offset, row in enumerate(rows):
        r_idx = data_start + row_offset
        row_fill = ALT_ROW_FILL if row_offset % 2 == 1 else None
        ws.row_dimensions[r_idx].height = 18

        for c_idx, col in enumerate(base_cols, start=1):
            val = get_row_value(row, col["key"], row_offset)
            cell = ws.cell(r_idx, c_idx, val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left" if col["key"] in ("item_name", "spec") else "center", vertical="center")
            cell.font = Font(size=10)
            if row_fill:
                cell.fill = row_fill

        c = len(base_cols) + 1

        if has_std:
            std_val = row.get("standard_unit_price") or ""
            cell = ws.cell(r_idx, c, int(to_number(std_val)) if std_val and to_number(std_val) else std_val)
            cell.border = BORDER
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(size=10)
            if row_fill:
                cell.fill = row_fill
            c += 1

        best_name, best_price = lowest_vendor(row, vendors)
        for v_idx, vendor in enumerate(vendors):
            price = get_vendor_value(row, vendor, "unit_price")
            amount = get_vendor_value(row, vendor, "amount")
            price_num = to_number(price) if price not in (None, "", "원문 미기재") else 0
            amount_num = to_number(amount) if amount not in (None, "", "원문 미기재") else 0

            is_lowest = best_name and comparable_company(vendor.get("name")) == comparable_company(best_name)
            cell_fill = GREEN_FILL if is_lowest and price_num else (row_fill or None)

            for val, col_offset, fmt in [(price_num or price, 0, "#,##0"), (amount_num or amount, 1, "#,##0")]:
                cell = ws.cell(r_idx, c + col_offset, int(to_number(val)) if val not in (None, "", "원문 미기재") and to_number(val) else val)
                cell.border = BORDER
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right" if to_number(val) else "center", vertical="center")
                cell.font = Font(size=10, bold=is_lowest and col_offset == 0)
                if cell_fill:
                    cell.fill = cell_fill
            c += 2

        # 최저업체/최저단가/비고
        lowest_cell = ws.cell(r_idx, c, best_name)
        lowest_cell.border = BORDER
        lowest_cell.alignment = Alignment(horizontal="center", vertical="center")
        lowest_cell.font = Font(size=10, bold=True, color="1A7A4A")
        if row_fill:
            lowest_cell.fill = row_fill

        price_cell = ws.cell(r_idx, c + 1, int(best_price) if best_price else "")
        price_cell.border = BORDER
        price_cell.number_format = "#,##0"
        price_cell.alignment = Alignment(horizontal="right", vertical="center")
        price_cell.font = Font(size=10, bold=True, color="1A7A4A")
        if row_fill:
            price_cell.fill = row_fill

        remark_val = get_row_value(row, "remark")
        remark_cell = ws.cell(r_idx, c + 2, remark_val)
        remark_cell.border = BORDER
        remark_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        remark_cell.font = Font(size=9, color="666666")
        if row_fill:
            remark_cell.fill = row_fill

    # 하단 빈 행 — 데이터 행과 동일한 테두리/높이 유지 (최소 5행, 최대 15행)
    empty_row_count = max(5, 15 - len(rows))
    for e_offset in range(empty_row_count):
        er = data_start + len(rows) + e_offset
        ws.row_dimensions[er].height = 18
        row_fill = ALT_ROW_FILL if (len(rows) + e_offset) % 2 == 1 else None
        for c_idx in range(1, total_cols + 1):
            cell = ws.cell(er, c_idx, "")
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=10)
            if row_fill:
                cell.fill = row_fill

    # 자동 필터 및 고정
    ws.auto_filter.ref = f"A{header_row + 1}:{get_column_letter(total_cols)}{data_start + len(rows) - 1}"
    ws.freeze_panes = ws.cell(data_start, 1)

    # A4 Landscape 출력 설정
    try:
        from openpyxl.worksheet.page import PageMargins
        ws.page_setup.paperSize = 9        # A4
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        ws.print_options.horizontalCentered = True
    except Exception:
        pass

    # 열 너비 최적화
    ws.column_dimensions["A"].width = 6   # NO
    col_letter_map = {col["key"]: get_column_letter(idx + 1) for idx, col in enumerate(base_cols)}
    if "construction_code" in col_letter_map:
        ws.column_dimensions[col_letter_map["construction_code"]].width = 16
    if "item_name" in col_letter_map:
        ws.column_dimensions[col_letter_map["item_name"]].width = 28
    if "spec" in col_letter_map:
        ws.column_dimensions[col_letter_map["spec"]].width = 18
    for c_i in range(len(base_cols) + std_extra + 1, total_cols - summary_cols + 1):
        ws.column_dimensions[get_column_letter(c_i)].width = 13
    ws.column_dimensions[get_column_letter(total_cols - 2)].width = 16  # 최저업체
    ws.column_dimensions[get_column_letter(total_cols - 1)].width = 13  # 최저단가
    ws.column_dimensions[get_column_letter(total_cols)].width = 20       # 비고

    return wb, {"template_kind": "ESTIMATE_COMPARISON", "vendor_count": len(vendors)}


def create_estimate_form_workbook(payload: Dict[str, Any]) -> Tuple[Workbook, Dict[str, Any]]:
    rows = as_list(payload.get("rows"))
    job = payload.get("job") or {}
    mapping = payload.get("mapping_json") or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"
    total_cols = 8
    merge_write(ws, 1, 1, 1, total_cols, mapping.get("title") or "견 적 서", bold=True, size=20, fill=TITLE_FILL)
    write_cell(ws, 3, 1, "견적일자", bold=True, fill=HEADER_FILL)
    write_cell(ws, 3, 2, today_text())
    write_cell(ws, 3, 3, "수신", bold=True, fill=HEADER_FILL)
    merge_write(ws, 3, 4, 3, 5, rows[0].get("recipient", "") if rows else "")
    write_cell(ws, 3, 6, "작성자", bold=True, fill=HEADER_FILL)
    merge_write(ws, 3, 7, 3, 8, payload.get("author_name") or "")
    write_cell(ws, 4, 1, "공급자", bold=True, fill=HEADER_FILL)
    merge_write(ws, 4, 2, 4, 4, rows[0].get("vendor_name", "") if rows else "")
    write_cell(ws, 4, 5, "견적명", bold=True, fill=HEADER_FILL)
    merge_write(ws, 4, 6, 4, 8, rows[0].get("document_title") or job.get("title") or "")
    headers = ["NO", "품명", "규격", "수량", "단위", "단가", "금액", "비고"]
    for idx, header in enumerate(headers, start=1):
        write_cell(ws, 6, idx, header, bold=True, fill=HEADER_FILL)
    total = 0
    for row_idx, row in enumerate(rows, start=7):
        values = [row_idx - 6, get_row_value(row, "item_name", row_idx-7), get_row_value(row, "spec"), get_row_value(row, "quantity"), get_row_value(row, "unit"), get_row_value(row, "unit_price"), get_row_value(row, "amount"), get_row_value(row, "remark")]
        total += to_number(values[6])
        for col_idx, value in enumerate(values, start=1):
            write_cell(ws, row_idx, col_idx, value, align="left" if col_idx in (2, 8) else "center")
            if col_idx in (6, 7):
                ws.cell(row_idx, col_idx).number_format = "#,##0"
    total_row = 7 + len(rows)
    merge_write(ws, total_row, 1, total_row, 6, "합계", bold=True, fill=LIGHT_FILL)
    write_cell(ws, total_row, 7, int(total) if total else "", bold=True, fill=LIGHT_FILL)
    ws.cell(total_row, 7).number_format = "#,##0"
    write_cell(ws, total_row, 8, "", fill=LIGHT_FILL)
    merge_write(ws, total_row + 2, 1, total_row + 2, 2, "특기사항", bold=True, fill=HEADER_FILL)
    merge_write(ws, total_row + 2, 3, total_row + 3, 8, rows[0].get("special_note", "") if rows else "", fill=TITLE_FILL)
    set_widths(ws, total_cols, 14)
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["H"].width = 24
    return wb, {"template_kind": "ESTIMATE_FORM"}

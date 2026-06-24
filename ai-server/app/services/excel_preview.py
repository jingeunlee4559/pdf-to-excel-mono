from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from app.services.storage_service import validate_storage_path


# 기본 indexed color 일부만 안전하게 보정합니다.
# 나머지는 잘못된 색 표시를 피하기 위해 fallback 처리합니다.
INDEXED_COLOR_FALLBACKS = {
    0: "#000000",
    1: "#FFFFFF",
    2: "#FF0000",
    3: "#00FF00",
    4: "#0000FF",
    5: "#FFFF00",
    6: "#FF00FF",
    7: "#00FFFF",
    8: "#000000",
    9: "#FFFFFF",
}


def _argb_to_hex(color: Any, fallback: str | None = None) -> str | None:
    try:
        if not color:
            return fallback

        color_type = getattr(color, "type", None)

        if color_type == "rgb" and getattr(color, "rgb", None):
            value = str(color.rgb)
            if value in {"00000000", "00FFFFFF"}:
                return fallback
            if len(value) == 8:
                return f"#{value[2:]}"
            if len(value) == 6:
                return f"#{value}"

        if color_type == "indexed":
            indexed = getattr(color, "indexed", None)
            return INDEXED_COLOR_FALLBACKS.get(indexed, fallback)

        # theme/tint 색상은 워크북 theme 해석이 필요합니다.
        # 여기서는 원본과 다른 색으로 왜곡하지 않도록 fallback을 사용합니다.
        if color_type == "theme":
            return fallback
    except Exception:
        return fallback
    return fallback


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _border_to_css(side: Any, fallback: str = "1px solid #e2e8f0") -> str:
    try:
        if not side or not side.style:
            return fallback
        color = _argb_to_hex(side.color, "#94a3b8") if side.color else "#94a3b8"
        # 브라우저 테이블에서 과도하게 두꺼워지는 것을 막기 위해 기본 1px로 표현합니다.
        return f"1px solid {color or '#94a3b8'}"
    except Exception:
        return fallback


def _cell_style(cell: Any) -> dict:
    fill = getattr(cell, "fill", None)
    font = getattr(cell, "font", None)
    alignment = getattr(cell, "alignment", None)
    border = getattr(cell, "border", None)

    bg = None
    if fill and getattr(fill, "fill_type", None) and getattr(fill, "fgColor", None):
        bg = _argb_to_hex(fill.fgColor, None)

    fg = "#0f172a"
    if font and getattr(font, "color", None):
        fg = _argb_to_hex(font.color, "#0f172a") or "#0f172a"

    horizontal = getattr(alignment, "horizontal", None) or "center"
    vertical = getattr(alignment, "vertical", None) or "center"

    return {
        "backgroundColor": bg or "#ffffff",
        "color": fg,
        "fontWeight": 700 if font and getattr(font, "bold", False) else 500,
        "fontSize": int(getattr(font, "sz", None) or 11) if font else 11,
        "fontFamily": getattr(font, "name", None) or "system-ui",
        "italic": bool(font and getattr(font, "italic", False)),
        "underline": bool(font and getattr(font, "underline", None)),
        "textAlign": horizontal,
        "verticalAlign": vertical,
        "whiteSpace": "normal" if alignment and getattr(alignment, "wrap_text", False) else "nowrap",
        "borderTop": _border_to_css(border.top) if border else "1px solid #e2e8f0",
        "borderRight": _border_to_css(border.right) if border else "1px solid #e2e8f0",
        "borderBottom": _border_to_css(border.bottom) if border else "1px solid #e2e8f0",
        "borderLeft": _border_to_css(border.left) if border else "1px solid #e2e8f0",
    }


def _build_merge_maps(ws: Any, max_row: int, max_col: int) -> tuple[dict[str, dict], dict[str, str], list[dict]]:
    """미리보기 범위 안의 병합 셀 정보를 만든다.

    openpyxl은 병합 영역의 좌상단 셀만 값/스타일을 갖고 나머지는 MergedCell로 취급한다.
    프론트에서 원본 양식처럼 보이려면 좌상단 셀에 rowSpan/colSpan을 주고 나머지 셀은 숨겨야 한다.
    """
    masters: dict[str, dict] = {}
    hidden: dict[str, str] = {}
    ranges: list[dict] = []

    for merged in ws.merged_cells.ranges:
        range_text = str(merged)
        min_col, min_row, max_merge_col, max_merge_row = range_boundaries(range_text)

        # 미리보기 범위와 전혀 겹치지 않으면 제외합니다.
        if min_row > max_row or min_col > max_col:
            continue

        clipped_max_row = min(max_merge_row, max_row)
        clipped_max_col = min(max_merge_col, max_col)
        master_address = f"{get_column_letter(min_col)}{min_row}"
        row_span = max(1, clipped_max_row - min_row + 1)
        col_span = max(1, clipped_max_col - min_col + 1)

        info = {
            "range": range_text,
            "masterAddress": master_address,
            "minRow": min_row,
            "minCol": min_col,
            "maxRow": max_merge_row,
            "maxCol": max_merge_col,
            "rowSpan": row_span,
            "colSpan": col_span,
        }
        masters[master_address] = info
        ranges.append(info)

        for row_idx in range(min_row, clipped_max_row + 1):
            for col_idx in range(min_col, clipped_max_col + 1):
                address = f"{get_column_letter(col_idx)}{row_idx}"
                if address != master_address:
                    hidden[address] = master_address

    return masters, hidden, ranges


def build_excel_preview(file_path: str, sheet_name: str | None = None, max_rows: int = 80, max_cols: int = 26) -> dict:
    from pathlib import Path as _Path
    _p = _Path(file_path).resolve()
    if _p.exists() and _p.is_file():
        target = _p
    else:
        target = validate_storage_path(file_path)
    if target.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("엑셀 미리보기는 xlsx, xlsm 파일만 지원합니다.")

    wb = load_workbook(target, read_only=False, data_only=False)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    max_row = min(ws.max_row or 1, max_rows)
    max_col = min(ws.max_column or 1, max_cols)

    columns = []
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        width = ws.column_dimensions[letter].width or 10
        hidden = bool(ws.column_dimensions[letter].hidden)
        columns.append({"index": col_idx, "letter": letter, "widthPx": max(48, int(width * 7)), "hidden": hidden})

    merge_masters, merge_hidden, merged_ranges = _build_merge_maps(ws, max_row, max_col)

    rows = []
    for row_idx in range(1, max_row + 1):
        height = ws.row_dimensions[row_idx].height or 24
        row_hidden = bool(ws.row_dimensions[row_idx].hidden)
        cells = []
        for col_idx in range(1, max_col + 1):
            address = f"{get_column_letter(col_idx)}{row_idx}"
            master_address = merge_hidden.get(address)

            if master_address:
                cells.append(
                    {
                        "address": address,
                        "row": row_idx,
                        "columnIndex": col_idx,
                        "columnLetter": get_column_letter(col_idx),
                        "text": "",
                        "value": "",
                        "isMergedHidden": True,
                        "masterAddress": master_address,
                    }
                )
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            merge_info = merge_masters.get(address)
            style_source = cell

            # 혹시 master 위치가 MergedCell이면 실제 좌상단 셀 스타일로 보정합니다.
            if isinstance(cell, MergedCell) and merge_info:
                style_source = ws.cell(row=merge_info["minRow"], column=merge_info["minCol"])

            cells.append(
                {
                    "address": cell.coordinate,
                    "row": row_idx,
                    "columnIndex": col_idx,
                    "columnLetter": get_column_letter(col_idx),
                    "value": _cell_value(cell.value),
                    "text": _cell_value(cell.value),
                    "isMerged": bool(merge_info),
                    "isMergedHidden": False,
                    "rowSpan": merge_info["rowSpan"] if merge_info else 1,
                    "colSpan": merge_info["colSpan"] if merge_info else 1,
                    "mergedRange": merge_info["range"] if merge_info else None,
                    "style": _cell_style(style_source),
                }
            )
        rows.append({"rowNumber": row_idx, "heightPx": max(24, int(height * 1.35)), "hidden": row_hidden, "cells": cells})

    return {
        "engine": "openpyxl",
        "sheetNames": wb.sheetnames,
        "sheet_names": wb.sheetnames,
        "preview": {
            "fileName": target.name,
            "sheetName": ws.title,
            "rows": rows,
            "columns": columns,
            "mergedRanges": [item["range"] for item in merged_ranges],
            "mergedCells": merged_ranges,
            "maxRow": ws.max_row,
            "maxColumn": ws.max_column,
        },
    }
